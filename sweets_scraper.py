"""
SWEETS Construction Website Scraper
Scrapes 3-Part Specifications, manufacturers, and product details from sweets.construction.com
Includes checkpoint/resume functionality for handling interruptions

Navigation path:
1. https://sweets.construction.com/quicklinks/3partspecs (All Divisions)
2. → Division page (e.g., /quicklinks/3partspecs/01-00-00-general-requirements)
3. → Section page (e.g., /masterformat/general-requirements-01-00-00/protecting-installed-construction-01-76-00)
4. → Product page (e.g., /manufacturer/prosoco-inc-nst154588/products/overcoat-...)

DOWNLOAD SECTION NOTES:
-----------------------
The Downloads section on SWEETS product pages (with tabs showing CAD (1), BIM (1), etc.)
is loaded dynamically via JavaScript using the loadDownloadTabs(selectedProductID) function.
This data is NOT present in the initial HTML response.

Two modes are available:
1. DEFAULT MODE (requests only): Detects content types from HTML indicators (Yes/No)
2. SELENIUM MODE (--use-selenium): Executes JavaScript to capture exact download counts

Use --use-selenium flag if you need exact download counts.
"""

import requests
from bs4 import BeautifulSoup
import time
import re
import json
import os
import signal
import sys
import atexit
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from dataclasses import dataclass, field
from typing import Optional, List, Dict
import logging
from datetime import datetime

# Optional Selenium imports for JavaScript-rendered content
try:
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from webdriver_manager.chrome import ChromeDriverManager
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Constants
BASE_URL = "https://sweets.construction.com"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
}
REQUEST_DELAY = 1.5  # Delay between requests

# Checkpoint settings
CHECKPOINT_DIR = "c:/Users/user1/Desktop/PersonalGit/Scraper/checkpoints"
CHECKPOINT_FILE = os.path.join(CHECKPOINT_DIR, "sweets_checkpoint.json")
CHECKPOINT_INTERVAL = 10  # Save checkpoint every N products scraped

# Retry settings
MAX_RETRIES = 3
RETRY_DELAY_BASE = 5
REQUEST_TIMEOUT = 30

# Output files
OUTPUT_FILE = "c:/Users/user1/Desktop/PersonalGit/Scraper/SWEETS_Scraped_Data.xlsx"
PARTIAL_SAVE_FILE = "c:/Users/user1/Desktop/PersonalGit/Scraper/SWEETS_Scraped_Data_Partial.xlsx"

# US State abbreviations to full names
STATE_ABBREV_TO_FULL = {
    'AL': 'Alabama', 'AK': 'Alaska', 'AZ': 'Arizona', 'AR': 'Arkansas',
    'CA': 'California', 'CO': 'Colorado', 'CT': 'Connecticut', 'DE': 'Delaware',
    'FL': 'Florida', 'GA': 'Georgia', 'HI': 'Hawaii', 'ID': 'Idaho',
    'IL': 'Illinois', 'IN': 'Indiana', 'IA': 'Iowa', 'KS': 'Kansas',
    'KY': 'Kentucky', 'LA': 'Louisiana', 'ME': 'Maine', 'MD': 'Maryland',
    'MA': 'Massachusetts', 'MI': 'Michigan', 'MN': 'Minnesota', 'MS': 'Mississippi',
    'MO': 'Missouri', 'MT': 'Montana', 'NE': 'Nebraska', 'NV': 'Nevada',
    'NH': 'New Hampshire', 'NJ': 'New Jersey', 'NM': 'New Mexico', 'NY': 'New York',
    'NC': 'North Carolina', 'ND': 'North Dakota', 'OH': 'Ohio', 'OK': 'Oklahoma',
    'OR': 'Oregon', 'PA': 'Pennsylvania', 'RI': 'Rhode Island', 'SC': 'South Carolina',
    'SD': 'South Dakota', 'TN': 'Tennessee', 'TX': 'Texas', 'UT': 'Utah',
    'VT': 'Vermont', 'VA': 'Virginia', 'WA': 'Washington', 'WV': 'West Virginia',
    'WI': 'Wisconsin', 'WY': 'Wyoming', 'DC': 'District of Columbia',
    # Canadian provinces
    'AB': 'Alberta', 'BC': 'British Columbia', 'MB': 'Manitoba', 'NB': 'New Brunswick',
    'NL': 'Newfoundland and Labrador', 'NS': 'Nova Scotia', 'ON': 'Ontario',
    'PE': 'Prince Edward Island', 'QC': 'Quebec', 'SK': 'Saskatchewan',
}


@dataclass
class Product:
    """Represents a product/company from SWEETS"""
    name: str
    url: str
    manufacturer_name: str = ""
    manufacturer_id: str = ""
    address: str = ""
    city: str = ""
    state: str = ""
    zip_code: str = ""
    phone: str = ""
    fax: str = ""
    email: str = ""
    website: str = ""
    division_code: str = ""
    division_name: str = ""
    section_code: str = ""
    section_name: str = ""
    category: str = ""
    masterformat: str = ""
    # Download/content types (Yes/No detection from HTML)
    has_3part_spec: bool = False
    has_bim: bool = False
    has_cad: bool = False
    has_ceu: bool = False
    has_catalog: bool = False
    has_data_sheet: bool = False
    has_gallery: bool = False
    has_green: bool = False
    has_product_selector: bool = False
    has_supporting_material: bool = False
    # Download counts (populated only with Selenium mode)
    count_3part_spec: int = 0
    count_bim: int = 0
    count_cad: int = 0
    count_ceu: int = 0
    count_catalog: int = 0
    count_gallery: int = 0
    count_green: int = 0
    count_other: int = 0
    count_total: int = 0
    description: str = ""


@dataclass
class Section:
    """Represents a MasterFormat section"""
    code: str
    name: str
    url: str
    item_count: int = 0
    products: List[Product] = field(default_factory=list)


@dataclass
class Division:
    """Represents a CSI Division"""
    code: str
    name: str
    url: str
    item_count: int = 0
    sections: List[Section] = field(default_factory=list)


class ProgressTracker:
    """Tracks scraping progress and provides ETA estimates"""

    def __init__(self):
        self.start_time = None
        self.total_items = 0
        self.scraped_items = 0
        self.scrape_times = []
        self.last_update_time = None

    def start(self, total_items: int = 0):
        self.start_time = time.time()
        self.total_items = total_items
        self.scraped_items = 0
        self.scrape_times = []
        self.last_update_time = time.time()

    def set_total(self, total: int):
        self.total_items = total

    def update(self, count: int = 1):
        current_time = time.time()
        if self.last_update_time:
            self.scrape_times.append(current_time - self.last_update_time)
            if len(self.scrape_times) > 100:
                self.scrape_times = self.scrape_times[-100:]
        self.last_update_time = current_time
        self.scraped_items += count

    def get_percentage(self) -> float:
        if self.total_items == 0:
            return 0.0
        return (self.scraped_items / self.total_items) * 100

    def get_eta_seconds(self) -> float:
        if not self.scrape_times or self.scraped_items == 0:
            return 0
        avg_time = sum(self.scrape_times) / len(self.scrape_times)
        remaining = self.total_items - self.scraped_items
        return avg_time * remaining

    def get_eta_formatted(self) -> str:
        eta_seconds = self.get_eta_seconds()
        if eta_seconds <= 0:
            return "Calculating..."
        hours = int(eta_seconds // 3600)
        minutes = int((eta_seconds % 3600) // 60)
        seconds = int(eta_seconds % 60)
        if hours > 0:
            return f"{hours}h {minutes}m {seconds}s"
        elif minutes > 0:
            return f"{minutes}m {seconds}s"
        return f"{seconds}s"

    def get_elapsed_formatted(self) -> str:
        if not self.start_time:
            return "0s"
        elapsed = time.time() - self.start_time
        hours = int(elapsed // 3600)
        minutes = int((elapsed % 3600) // 60)
        seconds = int(elapsed % 60)
        if hours > 0:
            return f"{hours}h {minutes}m {seconds}s"
        elif minutes > 0:
            return f"{minutes}m {seconds}s"
        return f"{seconds}s"

    def get_speed(self) -> float:
        if not self.scrape_times:
            return 0.0
        avg_time = sum(self.scrape_times) / len(self.scrape_times)
        if avg_time == 0:
            return 0.0
        return 60.0 / avg_time

    def get_progress_bar(self, width: int = 30) -> str:
        percentage = self.get_percentage()
        filled = int(width * percentage / 100)
        bar = "█" * filled + "░" * (width - filled)
        return f"[{bar}] {percentage:.1f}%"

    def get_status_line(self) -> str:
        return (
            f"{self.get_progress_bar()} | "
            f"{self.scraped_items:,}/{self.total_items:,} products | "
            f"Speed: {self.get_speed():.1f}/min | "
            f"ETA: {self.get_eta_formatted()} | "
            f"Elapsed: {self.get_elapsed_formatted()}"
        )


class SWEETSScraper:
    """Main scraper class for SWEETS website

    Args:
        use_selenium: If True, uses Selenium to capture exact download counts
                     from the dynamically-loaded Downloads section.
                     Requires selenium and webdriver-manager packages.
    """

    def __init__(self, use_selenium: bool = False):
        self.session = requests.Session()
        self.session.headers.update(HEADERS)
        self.divisions: List[Division] = []
        self.all_products: List[Product] = []
        self.products_scraped_count = 0
        self.interrupted = False
        self.progress = ProgressTracker()
        self.use_selenium = use_selenium
        self.driver = None
        self._manufacturer_cache: Dict[str, Dict[str, str]] = {}  # Cache manufacturer contact data

        self._setup_checkpoint_dir()
        self._setup_signal_handlers()

        if use_selenium:
            self._init_selenium()

    def _setup_checkpoint_dir(self):
        os.makedirs(CHECKPOINT_DIR, exist_ok=True)

    def _setup_signal_handlers(self):
        def signal_handler(signum, frame):
            logger.warning(f"\nInterrupt received (signal {signum}). Saving checkpoint and partial data...")
            self.interrupted = True
            self._save_checkpoint()
            self._save_partial_excel()
            logger.info("Checkpoint and partial data saved. Resume with --resume flag.")
            sys.exit(0)

        signal.signal(signal.SIGINT, signal_handler)
        signal.signal(signal.SIGTERM, signal_handler)
        atexit.register(self._emergency_save)

    def _emergency_save(self):
        if self.products_scraped_count > 0 and not self.interrupted:
            logger.warning("Emergency checkpoint save triggered")
            self._save_checkpoint()
            self._save_partial_excel()

    def _init_selenium(self):
        """Initialize Selenium WebDriver for JavaScript rendering"""
        if not SELENIUM_AVAILABLE:
            logger.error("Selenium not available. Install with: pip install selenium webdriver-manager")
            raise ImportError("selenium and webdriver-manager are required for --use-selenium mode")

        logger.info("Initializing Selenium WebDriver...")
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument(f"user-agent={HEADERS['User-Agent']}")

        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=chrome_options)
        logger.info("Selenium WebDriver initialized")

    def close(self):
        """Clean up resources"""
        if self.driver:
            self.driver.quit()
            self.driver = None

    def _extract_download_counts_selenium(self, product: Product) -> Product:
        """Extract exact download counts using Selenium to execute JavaScript.

        The Downloads section tabs show counts like: ALL (13), CAD (1), BIM (1), etc.
        This data is loaded via the loadDownloadTabs() JavaScript function.
        """
        if not self.driver:
            return product

        try:
            self.driver.get(product.url)

            # Wait for page to load and click the download button to trigger AJAX
            wait = WebDriverWait(self.driver, 10)

            # Try to find and click the download button to load the tabs
            try:
                download_btn = wait.until(
                    EC.element_to_be_clickable((By.ID, "downloadbtn"))
                )
                download_btn.click()
                time.sleep(2)  # Wait for AJAX to load download tabs
            except Exception:
                # Try alternative button
                try:
                    download_btn = self.driver.find_element(By.ID, "downloadbtnProductDesc")
                    download_btn.click()
                    time.sleep(2)
                except Exception:
                    pass

            # Now extract the download tab counts
            # Look for tab elements with counts like "CAD (1)", "BIM (1)"
            page_source = self.driver.page_source

            # Pattern to match tab text with counts: "TYPE (N)"
            tab_patterns = {
                'all': r'ALL\s*\((\d+)\)',
                'catalog': r'CATALOG[S]?\s*\((\d+)\)',
                'cad': r'CAD\s*\((\d+)\)',
                'bim': r'BIM\s*\((\d+)\)',
                '3part_spec': r'3-PART\s*SPEC[S]?\s*\((\d+)\)',
                'gallery': r'GALLER(?:Y|IES)\s*\((\d+)\)',
                'ceu': r'CEU\s*\((\d+)\)',
                'green': r'GREEN\s*\((\d+)\)',
                'other': r'OTHER\s*\((\d+)\)',
            }

            for key, pattern in tab_patterns.items():
                match = re.search(pattern, page_source, re.IGNORECASE)
                if match:
                    count = int(match.group(1))
                    if key == 'all':
                        product.count_total = count
                    elif key == 'catalog':
                        product.count_catalog = count
                        product.has_catalog = count > 0
                    elif key == 'cad':
                        product.count_cad = count
                        product.has_cad = count > 0
                    elif key == 'bim':
                        product.count_bim = count
                        product.has_bim = count > 0
                    elif key == '3part_spec':
                        product.count_3part_spec = count
                        product.has_3part_spec = count > 0
                    elif key == 'gallery':
                        product.count_gallery = count
                        product.has_gallery = count > 0
                    elif key == 'ceu':
                        product.count_ceu = count
                        product.has_ceu = count > 0
                    elif key == 'green':
                        product.count_green = count
                        product.has_green = count > 0
                    elif key == 'other':
                        product.count_other = count

            logger.debug(f"Download counts: Total={product.count_total}, CAD={product.count_cad}, "
                        f"BIM={product.count_bim}, 3-Part={product.count_3part_spec}")

        except Exception as e:
            logger.warning(f"Failed to extract download counts via Selenium: {e}")

        return product

    def _save_partial_excel(self):
        try:
            self.export_to_excel(PARTIAL_SAVE_FILE)
            logger.info(f"Partial data saved to {PARTIAL_SAVE_FILE}")
        except Exception as e:
            logger.error(f"Failed to save partial Excel: {e}")

    def _make_request(self, url: str) -> Optional[BeautifulSoup]:
        """Make HTTP request with retry logic"""
        for attempt in range(MAX_RETRIES):
            try:
                time.sleep(REQUEST_DELAY)
                response = self.session.get(url, timeout=REQUEST_TIMEOUT)
                response.raise_for_status()
                return BeautifulSoup(response.text, 'html.parser')
            except requests.exceptions.Timeout as e:
                retry_delay = RETRY_DELAY_BASE * (2 ** attempt)
                logger.warning(f"Timeout on {url} (attempt {attempt + 1}/{MAX_RETRIES}). Retrying in {retry_delay}s...")
                if attempt < MAX_RETRIES - 1:
                    time.sleep(retry_delay)
                else:
                    logger.error(f"Failed to fetch {url} after {MAX_RETRIES} attempts: {e}")
                    return None
            except requests.exceptions.ConnectionError as e:
                retry_delay = RETRY_DELAY_BASE * (2 ** attempt)
                logger.warning(f"Connection error on {url} (attempt {attempt + 1}/{MAX_RETRIES}). Retrying in {retry_delay}s...")
                if attempt < MAX_RETRIES - 1:
                    time.sleep(retry_delay)
                else:
                    logger.error(f"Failed to fetch {url} after {MAX_RETRIES} attempts: {e}")
                    return None
            except requests.RequestException as e:
                logger.error(f"Failed to fetch {url}: {e}")
                return None
        return None

    def _make_request_raw(self, url: str) -> Optional[str]:
        """Make HTTP request and return raw HTML"""
        for attempt in range(MAX_RETRIES):
            try:
                time.sleep(REQUEST_DELAY)
                response = self.session.get(url, timeout=REQUEST_TIMEOUT)
                response.raise_for_status()
                return response.text
            except requests.exceptions.Timeout as e:
                retry_delay = RETRY_DELAY_BASE * (2 ** attempt)
                logger.warning(f"Timeout on {url} (attempt {attempt + 1}/{MAX_RETRIES}). Retrying in {retry_delay}s...")
                if attempt < MAX_RETRIES - 1:
                    time.sleep(retry_delay)
                else:
                    logger.error(f"Failed to fetch {url} after {MAX_RETRIES} attempts: {e}")
                    return None
            except requests.RequestException as e:
                logger.error(f"Failed to fetch {url}: {e}")
                return None
        return None

    def scrape_divisions(self) -> List[Division]:
        """Scrape all divisions from the 3partspecs page"""
        url = f"{BASE_URL}/quicklinks/3partspecs"
        logger.info(f"Scraping divisions from {url}")

        soup = self._make_request(url)
        if not soup:
            return []

        divisions = []

        # Find division links matching pattern: /quicklinks/3partspecs/XX-00-00-name
        division_links = soup.find_all('a', href=re.compile(r'/quicklinks/3partspecs/\d+-\d+-\d+'))

        seen_urls = set()
        for link in division_links:
            href = link.get('href', '')
            if href in seen_urls:
                continue
            seen_urls.add(href)

            text = link.get_text(strip=True)
            full_url = f"{BASE_URL}{href}"

            # Parse code and name from text like "01 00 00 - General Requirements"
            match = re.match(r'^(\d+\s+\d+\s+\d+)\s*-\s*(.+)$', text)
            if match:
                code = match.group(1)
                name = match.group(2).strip()

                # Try to find item count (in the next sibling or nearby)
                item_count = 0
                parent = link.find_parent('tr') or link.find_parent('div')
                if parent:
                    count_match = re.search(r'(\d+)\s*$', parent.get_text())
                    if count_match:
                        item_count = int(count_match.group(1))

                division = Division(
                    code=code,
                    name=name,
                    url=full_url,
                    item_count=item_count
                )
                divisions.append(division)
                logger.info(f"Found division: {code} - {name} ({item_count} items)")

        self.divisions = divisions
        return divisions

    def scrape_division_sections(self, division: Division) -> Division:
        """Scrape sections from a division page"""
        logger.info(f"Scraping sections for division: {division.code} - {division.name}")

        soup = self._make_request(division.url)
        if not soup:
            return division

        # Find section links matching pattern: /masterformat/xxx/xxx
        # Include dots in the URL pattern to catch sub-sections like:
        #   /masterformat/masonry-04-00-00/unit-masonry-stabilization-04-01-20.41
        #   /masterformat/masonry-04-00-00/granite--04-40-00.17
        section_links = soup.find_all('a', href=re.compile(r'/masterformat/[\w-]+/[\w.%-]+'))

        seen_urls = set()
        for link in section_links:
            href = link.get('href', '')
            if href in seen_urls:
                continue
            seen_urls.add(href)

            text = link.get_text(strip=True)
            full_url = f"{BASE_URL}{href}"

            # Parse code and name from text like:
            #   "01 76 00 - Protecting Installed Construction"
            #   "04 01 20 - 04 01 20 - Maintenance of Unit Masonry"
            #   "04 05 23.19 - Masonry Cavity Drainage Weepholes And Vents"
            #   "04 40 00.17 - Granite"
            # Support dotted sub-codes like "04 01 20.41"
            match = re.match(
                r'^(\d+\s+\d+\s+\d+(?:\.\d+)?)\s*-\s*(?:\d+\s+\d+\s+\d+(?:\.\d+)?\s*-\s*)?(.+)$',
                text
            )
            if match:
                code = match.group(1)
                name = match.group(2).strip()

                # Try to find item count
                item_count = 0
                parent = link.find_parent('tr') or link.find_parent('div')
                if parent:
                    count_match = re.search(r'(\d+)\s*$', parent.get_text())
                    if count_match:
                        item_count = int(count_match.group(1))

                section = Section(
                    code=code,
                    name=name,
                    url=full_url,
                    item_count=item_count
                )
                division.sections.append(section)
                logger.info(f"  Found section: {code} - {name} ({item_count} items)")

        logger.info(f"Found {len(division.sections)} sections in division {division.code}")
        return division

    def scrape_section_products(self, section: Section, division: Division) -> Section:
        """Scrape products from a section page"""
        logger.info(f"Scraping products for section: {section.code} - {section.name}")

        soup = self._make_request(section.url)
        if not soup:
            return section

        # Find product links matching pattern: /manufacturer/xxx/products/xxx
        product_links = soup.find_all('a', href=re.compile(r'/manufacturer/[\w-]+/products/[\w-]+'))

        seen_urls = set()
        for link in product_links:
            href = link.get('href', '')

            # Skip empty hrefs or duplicates
            base_href = href.split('?')[0]  # Remove query params
            if not href or base_href in seen_urls:
                continue

            text = link.get_text(strip=True)
            if not text:  # Skip links without text (usually image links)
                continue

            seen_urls.add(base_href)
            full_url = f"{BASE_URL}{href}"

            # Parse manufacturer and product name from text like "PROSOCO Inc. - Product Name"
            manufacturer_name = ""
            product_name = text
            if " - " in text:
                parts = text.split(" - ", 1)
                manufacturer_name = parts[0].strip()
                product_name = parts[1].strip() if len(parts) > 1 else text

            # Extract manufacturer ID from URL
            id_match = re.search(r'/manufacturer/([\w-]+)/products/', href)
            manufacturer_id = id_match.group(1) if id_match else ""

            product = Product(
                name=product_name,
                url=full_url,
                manufacturer_name=manufacturer_name,
                manufacturer_id=manufacturer_id,
                division_code=division.code,
                division_name=division.name,
                section_code=section.code,
                section_name=section.name
            )
            section.products.append(product)

        logger.info(f"  Found {len(section.products)} products in section {section.code}")
        return section

    def _scrape_manufacturer_contact(self, manufacturer_id: str) -> Dict[str, str]:
        """Fallback: scrape the manufacturer's main page for contact info.

        URL pattern: /manufacturer/{manufacturer-id}
        """
        mfr_url = f"{BASE_URL}/manufacturer/{manufacturer_id}"
        logger.debug(f"Fetching manufacturer page for missing address: {mfr_url}")

        try:
            html = self._make_request_raw(mfr_url)
            if not html:
                return {}

            soup = BeautifulSoup(html, 'html.parser')
            company_info_div = soup.find('div', class_='companyInfo')
            if company_info_div:
                address_tag = company_info_div.find('address')
                if address_tag:
                    return self._parse_address_tag(address_tag)
        except Exception as e:
            logger.debug(f"Failed to scrape manufacturer page {mfr_url}: {e}")

        return {}

    def _parse_address_tag(self, address_tag, manufacturer_name: str = "") -> Dict[str, str]:
        """Parse the <address> tag from companyInfo div to extract contact details.

        The <address> tag has a consistent structure with <br>-separated lines:
            <br>Company Name
            <br>Street Line 1
            <br>Street Line 2 (optional, e.g. Suite/P.O. Box)
            <br>City, ST ZIP
            <br>Tel: (XXX) XXX-XXXX
            <br>Fax: (XXX) XXX-XXXX
            <p><a href="mailto:email">email</a></p>
            <p><a href="https://website">website</a></p>
        """
        result = {
            'address': '', 'city': '', 'state': '', 'zip_code': '',
            'phone': '', 'fax': '', 'email': '', 'website': ''
        }

        if not address_tag:
            return result

        # Extract email from mailto link within the address tag
        excluded_email_domains = ['construction.com', 'sweets.com', 'sso.construction.com',
                                  'noreply', 'donotreply', 'unsubscribe']
        mailto_link = address_tag.find('a', href=re.compile(r'^mailto:', re.IGNORECASE))
        if mailto_link:
            email = re.sub(r'^mailto:', '', mailto_link.get('href', ''), flags=re.IGNORECASE).strip()
            if not any(exc in email.lower() for exc in excluded_email_domains):
                result['email'] = email

        # Extract website from links within the address tag
        excluded_domains = ['construction.com', 'sweets.com', 'sso.construction.com',
                          'facebook.com', 'linkedin.com', 'twitter.com', 'instagram.com',
                          'youtube.com', 'google.com', 'pinterest.com', 'tiktok.com',
                          'amazonaws.com', 'cloudfront.net', 'cloudflare.com',
                          'googleapis.com', 'gstatic.com', 'bootstrapcdn.com', 'jquery.com',
                          'fontawesome.com', 'microsoft.com', 'bing.com', 'w3.org', 'schema.org',
                          'doubleclick.net', 'googlesyndication.com', 'googletagmanager.com']

        for link in address_tag.find_all('a', href=re.compile(r'^https?://')):
            href = link.get('href', '')
            # Skip mailto, javascript, and anchor links
            if href.startswith('mailto:') or href.startswith('javascript:'):
                continue
            domain_match = re.search(r'https?://(?:www\.)?([a-zA-Z0-9.-]+)', href)
            if domain_match:
                domain_str = domain_match.group(1).lower()
                if not any(exc in domain_str for exc in excluded_domains):
                    if not any(ext in href.lower() for ext in ['.js', '.css', '.png', '.jpg', '.gif', '.ico']):
                        result['website'] = href.rstrip('/')
                        break

        # Get text content split by <br> tags
        # Replace <br> tags with a delimiter, then split
        address_html = str(address_tag)
        # Replace <br>, <br/>, <br /> with delimiter
        address_html = re.sub(r'<br\s*/?\s*>', '\n', address_html, flags=re.IGNORECASE)
        # Remove all other HTML tags
        address_text = re.sub(r'<[^>]+>', '', address_html)
        lines = [line.strip() for line in address_text.split('\n') if line.strip()]

        # Parse lines sequentially
        address_lines = []
        city_state_found = False

        for line in lines:
            # Skip company name (first non-empty line, matches manufacturer)
            # Skip "Find a rep" and "Request More Info" type lines
            if any(skip in line.lower() for skip in ['find a rep', 'request more info', 'request info']):
                continue

            # Check for Tel: line
            tel_match = re.match(r'Tel:\s*(.+)', line, re.IGNORECASE)
            if tel_match:
                phone_raw = tel_match.group(1).strip()
                # Normalize phone format
                phone_digits = re.match(r'\((\d{3})\)\s*(\d{3})-(\d{4})', phone_raw)
                if phone_digits:
                    result['phone'] = f"({phone_digits.group(1)}) {phone_digits.group(2)}-{phone_digits.group(3)}"
                else:
                    phone_digits2 = re.match(r'(\d{3})[-.\s](\d{3})[-.\s](\d{4})', phone_raw)
                    if phone_digits2:
                        result['phone'] = f"({phone_digits2.group(1)}) {phone_digits2.group(2)}-{phone_digits2.group(3)}"
                    else:
                        result['phone'] = phone_raw
                continue

            # Check for Fax: line
            fax_match = re.match(r'Fax:\s*(.+)', line, re.IGNORECASE)
            if fax_match:
                fax_raw = fax_match.group(1).strip()
                fax_digits = re.match(r'\((\d{3})\)\s*(\d{3})-(\d{4})', fax_raw)
                if fax_digits:
                    result['fax'] = f"({fax_digits.group(1)}) {fax_digits.group(2)}-{fax_digits.group(3)}"
                else:
                    fax_digits2 = re.match(r'(\d{3})[-.\s](\d{3})[-.\s](\d{4})', fax_raw)
                    if fax_digits2:
                        result['fax'] = f"({fax_digits2.group(1)}) {fax_digits2.group(2)}-{fax_digits2.group(3)}"
                    else:
                        result['fax'] = fax_raw
                continue

            # Skip email/website text lines (already extracted from links)
            if '@' in line or line.startswith('http'):
                continue

            if city_state_found:
                continue

            # Check for "City, ST ZIP" pattern (US)
            city_state_zip = re.match(
                r'^([A-Za-z][A-Za-z\s.]+?),\s*([A-Z]{2})\s+(\d{5}(?:-\d{4})?)$', line
            )
            if city_state_zip:
                result['city'] = city_state_zip.group(1).strip()
                state_abbr = city_state_zip.group(2).strip()
                result['state'] = STATE_ABBREV_TO_FULL.get(state_abbr, state_abbr)
                result['zip_code'] = city_state_zip.group(3).strip()
                city_state_found = True
                continue

            # Check for Canadian "City, Province PostalCode" pattern
            city_state_ca = re.match(
                r'^([A-Za-z][A-Za-z\s.]+?),\s*([A-Z]{2})\s+([A-Z]\d[A-Z]\s*\d[A-Z]\d)$', line
            )
            if city_state_ca:
                result['city'] = city_state_ca.group(1).strip()
                state_abbr = city_state_ca.group(2).strip()
                result['state'] = STATE_ABBREV_TO_FULL.get(state_abbr, state_abbr)
                result['zip_code'] = city_state_ca.group(3).strip()
                city_state_found = True
                continue

            # Check for "City, Province/State PostalCode" with full province name
            city_prov_ca = re.match(
                r'^([A-Za-z][A-Za-z\s.]+?),\s*([A-Za-z\s]+?)\s+([A-Z]\d[A-Z]\s*\d[A-Z]\d)$', line
            )
            if city_prov_ca:
                result['city'] = city_prov_ca.group(1).strip()
                result['state'] = city_prov_ca.group(2).strip()
                result['zip_code'] = city_prov_ca.group(3).strip()
                city_state_found = True
                continue

            # Otherwise, this is an address line (street, P.O. Box, suite, etc.)
            address_lines.append(line)

        # Build address from collected lines (skip the first line if it matches manufacturer name)
        if address_lines:
            # The first line might be the company name - skip it if it matches
            if manufacturer_name and address_lines[0].lower().replace(',', '').replace('.', '') == \
               manufacturer_name.lower().replace(',', '').replace('.', ''):
                address_lines = address_lines[1:]

            result['address'] = ', '.join(address_lines)

        return result

    def scrape_product_details(self, product: Product) -> Product:
        """Scrape detailed product/company information from product page"""
        logger.debug(f"Scraping product: {product.manufacturer_name} - {product.name}")

        html = self._make_request_raw(product.url)
        if not html:
            return product

        soup = BeautifulSoup(html, 'html.parser')

        # ===== PRIMARY: Parse <address> tag from companyInfo div =====
        company_info_div = soup.find('div', class_='companyInfo')
        address_tag = None
        if company_info_div:
            address_tag = company_info_div.find('address')

        if address_tag:
            contact = self._parse_address_tag(address_tag, product.manufacturer_name)
            product.phone = contact['phone']
            product.fax = contact['fax']
            product.email = contact['email']
            product.website = contact['website']
            product.address = contact['address']
            product.city = contact['city']
            product.state = contact['state']
            product.zip_code = contact['zip_code']
        else:
            # ===== FALLBACK: Regex-based extraction if no <address> tag =====
            logger.debug(f"No <address> tag found for {product.manufacturer_name}, using regex fallback")

            # Phone
            tel_match = re.search(r'Tel:\s*\((\d{3})\)\s*(\d{3})-(\d{4})', html)
            if tel_match:
                product.phone = f"({tel_match.group(1)}) {tel_match.group(2)}-{tel_match.group(3)}"
            else:
                tel_match2 = re.search(r'Tel:\s*(\d{3})[-.\s](\d{3})[-.\s](\d{4})', html)
                if tel_match2:
                    product.phone = f"({tel_match2.group(1)}) {tel_match2.group(2)}-{tel_match2.group(3)}"

            # Fax
            fax_match = re.search(r'Fax:\s*\((\d{3})\)\s*(\d{3})-(\d{4})', html)
            if fax_match:
                product.fax = f"({fax_match.group(1)}) {fax_match.group(2)}-{fax_match.group(3)}"

            # Email
            email_matches = re.findall(r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})', html)
            excluded_email_domains = ['construction.com', 'sweets.com', 'noreply', 'donotreply', 'unsubscribe']
            for email in email_matches:
                if not any(exc in email.lower() for exc in excluded_email_domains):
                    product.email = email
                    break

            # Website
            excluded_domains = ['construction.com', 'sweets.com', 'facebook.com', 'linkedin.com', 'twitter.com',
                              'instagram.com', 'youtube.com', 'google.com', 'pinterest.com', 'tiktok.com',
                              'amazonaws.com', 'cloudfront.net', 'cloudflare.com',
                              'googleapis.com', 'gstatic.com', 'bootstrapcdn.com', 'jquery.com',
                              'fontawesome.com', 'microsoft.com', 'bing.com', 'w3.org', 'schema.org',
                              'doubleclick.net', 'googlesyndication.com', 'googletagmanager.com']
            website_links = re.findall(r'href="(https?://(?:www\.)?[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}/?)"', html)
            for url in website_links:
                domain = re.search(r'https?://(?:www\.)?([a-zA-Z0-9.-]+\.[a-zA-Z]{2,})', url)
                if domain:
                    domain_str = domain.group(1).lower()
                    if not any(exc in domain_str for exc in excluded_domains):
                        if not any(ext in url.lower() for ext in ['.js', '.css', '.png', '.jpg', '.gif', '.ico', '.woff', '.svg']):
                            product.website = url.rstrip('/')
                            break

            # Address (regex fallback)
            addr_pattern = re.search(
                r'>([^<]+)<[^>]*>([A-Za-z][A-Za-z\s.]+?),?\s*([A-Z]{2})\s+(\d{5}(?:-\d{4})?)<',
                html, re.IGNORECASE
            )
            if addr_pattern:
                product.address = addr_pattern.group(1).strip()
                product.city = addr_pattern.group(2).strip()
                state_abbr = addr_pattern.group(3).strip()
                product.state = STATE_ABBREV_TO_FULL.get(state_abbr, state_abbr)
                product.zip_code = addr_pattern.group(4).strip()

        # ===== MANUFACTURER PAGE FALLBACK =====
        # If address is still blank, try the manufacturer's main page
        if not product.address and product.manufacturer_id:
            cached = self._manufacturer_cache.get(product.manufacturer_id)
            if cached is not None:
                # Use cached data (could be empty dict if already tried and failed)
                if cached:
                    product.address = cached.get('address', '')
                    product.city = cached.get('city', '')
                    product.state = cached.get('state', '')
                    product.zip_code = cached.get('zip_code', '')
                    if not product.phone:
                        product.phone = cached.get('phone', '')
                    if not product.fax:
                        product.fax = cached.get('fax', '')
                    if not product.email:
                        product.email = cached.get('email', '')
                    if not product.website:
                        product.website = cached.get('website', '')
            else:
                mfr_contact = self._scrape_manufacturer_contact(product.manufacturer_id)
                self._manufacturer_cache[product.manufacturer_id] = mfr_contact
                if mfr_contact:
                    product.address = mfr_contact.get('address', '')
                    product.city = mfr_contact.get('city', '')
                    product.state = mfr_contact.get('state', '')
                    product.zip_code = mfr_contact.get('zip_code', '')
                    if not product.phone:
                        product.phone = mfr_contact.get('phone', '')
                    if not product.fax:
                        product.fax = mfr_contact.get('fax', '')
                    if not product.email:
                        product.email = mfr_contact.get('email', '')
                    if not product.website:
                        product.website = mfr_contact.get('website', '')

        # ===== CATEGORY/MASTERFORMAT =====
        category_match = re.search(r'Category:\s*</strong>\s*([^<]+)', html)
        if category_match:
            product.category = category_match.group(1).strip()
        else:
            category_match2 = re.search(r'Category:\s*([^<\n]+)', html)
            if category_match2:
                product.category = category_match2.group(1).strip()

        masterformat_match = re.search(r'MasterFormat:\s*</strong>\s*([^<]+)', html)
        if masterformat_match:
            product.masterformat = masterformat_match.group(1).strip()
        else:
            masterformat_match2 = re.search(r'MasterFormat:\s*([^<\n]+)', html)
            if masterformat_match2:
                product.masterformat = masterformat_match2.group(1).strip()

        # ===== DESCRIPTION =====
        desc_elem = soup.find('meta', attrs={'name': 'description'})
        if desc_elem:
            product.description = desc_elem.get('content', '')[:500]

        # ===== DOWNLOAD/CONTENT TYPES =====
        # IMPORTANT: On SWEETS, the Downloads section with tabs (ALL, CATALOGS, CAD, BIM, etc.)
        # and their counts (like "CAD (1)", "BIM (1)") are loaded via AJAX using the
        # loadDownloadTabs(selectedProductID) function. This data is NOT in the initial HTML.
        #
        # To capture exact counts, you would need to either:
        # 1. Use Selenium to execute JavaScript and wait for the downloads section to load
        # 2. Call the underlying API endpoint directly (if discovered)
        #
        # Current approach: Detect content type indicators from the HTML markup and metadata.
        # This gives Yes/No detection but NOT exact counts.

        html_lower = html.lower()

        # Look for download-related data in the page
        # Check product ID for download tracking
        product_id_match = re.search(r'selectedProductID\s*=\s*["\']?(\d+)', html)
        if product_id_match:
            product.manufacturer_id = product_id_match.group(1)

        # Detect content types from page indicators, meta tags, and embedded data
        # These patterns look for mentions in the page that indicate availability

        # 3-Part Spec / System Specifications
        product.has_3part_spec = bool(
            re.search(r'3-part\s*spec', html_lower) or
            re.search(r'3partspec', html_lower) or
            re.search(r'system\s*specifications', html_lower) or
            re.search(r'specwizard', html_lower) or
            re.search(r'"spec"', html_lower) or
            re.search(r'specification\s*document', html_lower) or
            # Look in data attributes or JSON
            re.search(r'downloadtype["\s:]+spec', html_lower)
        )

        # BIM - Revit/Building Information Modeling
        product.has_bim = bool(
            re.search(r'bim\s*library', html_lower) or
            re.search(r'"bim"', html_lower) or
            re.search(r'\.rfa', html_lower) or  # Revit family file
            re.search(r'\.rvt', html_lower) or  # Revit file
            re.search(r'revit\s*content', html_lower) or
            re.search(r'revit\s*family', html_lower) or
            re.search(r'bim\s*object', html_lower) or
            re.search(r'downloadtype["\s:]+bim', html_lower)
        )

        # CAD - AutoCAD files
        product.has_cad = bool(
            re.search(r'cad\s*library', html_lower) or
            re.search(r'"cad"', html_lower) or
            re.search(r'\.dwg', html_lower) or  # AutoCAD drawing
            re.search(r'\.dxf', html_lower) or  # Drawing exchange format
            re.search(r'autocad', html_lower) or
            re.search(r'cad\s*details', html_lower) or
            re.search(r'cad\s*drawing', html_lower) or
            re.search(r'downloadtype["\s:]+cad', html_lower)
        )

        # CEU (Continuing Education Units)
        product.has_ceu = bool(
            re.search(r'"ceu"', html_lower) or
            re.search(r'continuing\s*education', html_lower) or
            re.search(r'aia\s*credit', html_lower) or
            re.search(r'aia\s*ces', html_lower) or
            re.search(r'education\s*course', html_lower) or
            re.search(r'learning\s*unit', html_lower) or
            re.search(r'downloadtype["\s:]+ceu', html_lower)
        )

        # Catalog
        product.has_catalog = bool(
            re.search(r'"catalog"', html_lower) or
            re.search(r'product\s*catalog', html_lower) or
            re.search(r'full\s*catalog', html_lower) or
            re.search(r'catalog\s*download', html_lower) or
            re.search(r'downloadtype["\s:]+catalog', html_lower)
        )

        # Data Sheet / Technical Data
        product.has_data_sheet = bool(
            re.search(r'data\s*sheet', html_lower) or
            re.search(r'product\s*data', html_lower) or
            re.search(r'technical\s*data', html_lower) or
            re.search(r'tech\s*sheet', html_lower) or
            re.search(r'specification\s*sheet', html_lower) or
            re.search(r'downloadtype["\s:]+data', html_lower)
        )

        # Gallery / Case Studies
        product.has_gallery = bool(
            re.search(r'"gallery"', html_lower) or
            re.search(r'photo\s*gallery', html_lower) or
            re.search(r'image\s*gallery', html_lower) or
            re.search(r'case\s*stud', html_lower) or
            re.search(r'project\s*gallery', html_lower) or
            re.search(r'downloadtype["\s:]+gallery', html_lower)
        )

        # Green Building / Sustainability
        product.has_green = bool(
            re.search(r'"green"', html_lower) or
            re.search(r'green\s*building', html_lower) or
            re.search(r'leed\s*credit', html_lower) or
            re.search(r'leed\s*certif', html_lower) or
            re.search(r'sustainab', html_lower) or
            re.search(r'environmental\s*product', html_lower) or
            re.search(r'epd', html_lower) or  # Environmental Product Declaration
            re.search(r'downloadtype["\s:]+green', html_lower)
        )

        # Product Selector / Line Card
        product.has_product_selector = bool(
            re.search(r'product\s*selector', html_lower) or
            re.search(r'line\s*card', html_lower) or
            re.search(r'selection\s*guide', html_lower) or
            re.search(r'product\s*guide', html_lower) or
            re.search(r'downloadtype["\s:]+selector', html_lower)
        )

        # Supporting Material / Literature
        product.has_supporting_material = bool(
            re.search(r'supporting\s*material', html_lower) or
            re.search(r'literature', html_lower) or
            re.search(r'brochure', html_lower) or
            re.search(r'marketing\s*material', html_lower) or
            re.search(r'sales\s*sheet', html_lower) or
            re.search(r'downloadtype["\s:]+support', html_lower)
        )

        # If Selenium mode is enabled, extract exact download counts
        if self.use_selenium and self.driver:
            self._extract_download_counts_selenium(product)

        logger.info(f"Scraped: {product.manufacturer_name} | {product.city}, {product.state} | Phone: {product.phone} | Website: {product.website[:30] if product.website else 'N/A'}...")
        return product

    def _save_checkpoint(self):
        """Save current progress to checkpoint file"""
        checkpoint_data = {
            "timestamp": datetime.now().isoformat(),
            "products_scraped": self.products_scraped_count,
            "divisions": []
        }

        for div in self.divisions:
            div_data = {
                "code": div.code,
                "name": div.name,
                "url": div.url,
                "item_count": div.item_count,
                "sections": []
            }
            for section in div.sections:
                section_data = {
                    "code": section.code,
                    "name": section.name,
                    "url": section.url,
                    "item_count": section.item_count,
                    "products": []
                }
                for product in section.products:
                    section_data["products"].append({
                        "name": product.name,
                        "url": product.url,
                        "manufacturer_name": product.manufacturer_name,
                        "manufacturer_id": product.manufacturer_id,
                        "address": product.address,
                        "city": product.city,
                        "state": product.state,
                        "zip_code": product.zip_code,
                        "phone": product.phone,
                        "fax": product.fax,
                        "email": product.email,
                        "website": product.website,
                        "division_code": product.division_code,
                        "division_name": product.division_name,
                        "section_code": product.section_code,
                        "section_name": product.section_name,
                        "category": product.category,
                        "masterformat": product.masterformat,
                        "has_3part_spec": product.has_3part_spec,
                        "has_bim": product.has_bim,
                        "has_cad": product.has_cad,
                        "has_ceu": product.has_ceu,
                        "has_catalog": product.has_catalog,
                        "has_data_sheet": product.has_data_sheet,
                        "has_gallery": product.has_gallery,
                        "has_green": product.has_green,
                        "has_product_selector": product.has_product_selector,
                        "has_supporting_material": product.has_supporting_material,
                        "count_3part_spec": product.count_3part_spec,
                        "count_bim": product.count_bim,
                        "count_cad": product.count_cad,
                        "count_ceu": product.count_ceu,
                        "count_catalog": product.count_catalog,
                        "count_gallery": product.count_gallery,
                        "count_green": product.count_green,
                        "count_other": product.count_other,
                        "count_total": product.count_total,
                        "description": product.description
                    })
                div_data["sections"].append(section_data)
            checkpoint_data["divisions"].append(div_data)

        try:
            with open(CHECKPOINT_FILE, 'w', encoding='utf-8') as f:
                json.dump(checkpoint_data, f, indent=2, ensure_ascii=False)
            logger.info(f"Checkpoint saved: {self.products_scraped_count} products to {CHECKPOINT_FILE}")
        except Exception as e:
            logger.error(f"Failed to save checkpoint: {e}")

    def _load_checkpoint(self) -> dict:
        """Load checkpoint from file if it exists"""
        if not os.path.exists(CHECKPOINT_FILE):
            return None

        try:
            with open(CHECKPOINT_FILE, 'r', encoding='utf-8') as f:
                checkpoint_data = json.load(f)
            logger.info(f"Loaded checkpoint from {checkpoint_data['timestamp']}")
            logger.info(f"Previously scraped: {checkpoint_data['products_scraped']} products")
            return checkpoint_data
        except Exception as e:
            logger.error(f"Failed to load checkpoint: {e}")
            return None

    def _restore_from_checkpoint(self, checkpoint_data: dict):
        """Restore scraper state from checkpoint data"""
        for div_data in checkpoint_data.get("divisions", []):
            division = Division(
                code=div_data["code"],
                name=div_data["name"],
                url=div_data["url"],
                item_count=div_data.get("item_count", 0)
            )
            for section_data in div_data.get("sections", []):
                section = Section(
                    code=section_data["code"],
                    name=section_data["name"],
                    url=section_data["url"],
                    item_count=section_data.get("item_count", 0)
                )
                for prod_data in section_data.get("products", []):
                    product = Product(
                        name=prod_data["name"],
                        url=prod_data["url"],
                        manufacturer_name=prod_data.get("manufacturer_name", ""),
                        manufacturer_id=prod_data.get("manufacturer_id", ""),
                        address=prod_data.get("address", ""),
                        city=prod_data.get("city", ""),
                        state=prod_data.get("state", ""),
                        zip_code=prod_data.get("zip_code", ""),
                        phone=prod_data.get("phone", ""),
                        fax=prod_data.get("fax", ""),
                        email=prod_data.get("email", ""),
                        website=prod_data.get("website", ""),
                        division_code=prod_data.get("division_code", ""),
                        division_name=prod_data.get("division_name", ""),
                        section_code=prod_data.get("section_code", ""),
                        section_name=prod_data.get("section_name", ""),
                        category=prod_data.get("category", ""),
                        masterformat=prod_data.get("masterformat", ""),
                        has_3part_spec=prod_data.get("has_3part_spec", False),
                        has_bim=prod_data.get("has_bim", False),
                        has_cad=prod_data.get("has_cad", False),
                        has_ceu=prod_data.get("has_ceu", False),
                        has_catalog=prod_data.get("has_catalog", False),
                        has_data_sheet=prod_data.get("has_data_sheet", False),
                        has_gallery=prod_data.get("has_gallery", False),
                        has_green=prod_data.get("has_green", False),
                        has_product_selector=prod_data.get("has_product_selector", False),
                        has_supporting_material=prod_data.get("has_supporting_material", False),
                        count_3part_spec=prod_data.get("count_3part_spec", 0),
                        count_bim=prod_data.get("count_bim", 0),
                        count_cad=prod_data.get("count_cad", 0),
                        count_ceu=prod_data.get("count_ceu", 0),
                        count_catalog=prod_data.get("count_catalog", 0),
                        count_gallery=prod_data.get("count_gallery", 0),
                        count_green=prod_data.get("count_green", 0),
                        count_other=prod_data.get("count_other", 0),
                        count_total=prod_data.get("count_total", 0),
                        description=prod_data.get("description", "")
                    )
                    section.products.append(product)
                division.sections.append(section)
            self.divisions.append(division)

        self.products_scraped_count = checkpoint_data.get("products_scraped", 0)
        logger.info(f"Restored {len(self.divisions)} divisions from checkpoint")

    def _maybe_save_checkpoint(self):
        """Save checkpoint if enough products have been scraped since last save"""
        if self.products_scraped_count > 0 and self.products_scraped_count % CHECKPOINT_INTERVAL == 0:
            self._save_checkpoint()

    def clear_checkpoint(self):
        """Delete checkpoint file"""
        if os.path.exists(CHECKPOINT_FILE):
            os.remove(CHECKPOINT_FILE)
            logger.info("Checkpoint file cleared")

    def export_to_excel(self, filename: str = OUTPUT_FILE):
        """Export scraped data to Excel file"""
        logger.info(f"Exporting data to {filename}")

        wb = Workbook()
        ws = wb.active
        ws.title = "SWEETS Data"

        # Define headers - base columns
        headers = [
            "Division Code",
            "Division Name",
            "Section Code",
            "Section Name",
            "Manufacturer",
            "Product Name",
            "Address",
            "City",
            "State",
            "ZIP",
            "Phone",
            "Fax",
            "Email",
            "Website",
            "Category",
            "MasterFormat",
            "3-Part Spec",
            "BIM",
            "CAD",
            "CEU",
            "Catalog",
            "Data Sheet",
            "Gallery",
            "Green",
            "Product Selector",
            "Supporting Material",
        ]

        # Add download count columns if Selenium mode was used
        if self.use_selenium:
            headers.extend([
                "# 3-Part Spec",
                "# BIM",
                "# CAD",
                "# CEU",
                "# Catalog",
                "# Gallery",
                "# Green",
                "# Other",
                "# Total Downloads",
            ])

        headers.extend([
            "Product URL",
            "Source"
        ])

        # Style for headers
        header_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.border = thin_border

        # Write data
        row = 2
        for division in self.divisions:
            for section in division.sections:
                for product in section.products:
                    col = 1
                    ws.cell(row=row, column=col, value=product.division_code); col += 1
                    ws.cell(row=row, column=col, value=product.division_name); col += 1
                    ws.cell(row=row, column=col, value=product.section_code); col += 1
                    ws.cell(row=row, column=col, value=product.section_name); col += 1
                    ws.cell(row=row, column=col, value=product.manufacturer_name); col += 1
                    ws.cell(row=row, column=col, value=product.name); col += 1
                    ws.cell(row=row, column=col, value=product.address); col += 1
                    ws.cell(row=row, column=col, value=product.city); col += 1
                    ws.cell(row=row, column=col, value=product.state); col += 1
                    ws.cell(row=row, column=col, value=product.zip_code); col += 1
                    ws.cell(row=row, column=col, value=product.phone); col += 1
                    ws.cell(row=row, column=col, value=product.fax); col += 1
                    ws.cell(row=row, column=col, value=product.email); col += 1
                    ws.cell(row=row, column=col, value=product.website); col += 1
                    ws.cell(row=row, column=col, value=product.category); col += 1
                    ws.cell(row=row, column=col, value=product.masterformat); col += 1
                    ws.cell(row=row, column=col, value="Yes" if product.has_3part_spec else "No"); col += 1
                    ws.cell(row=row, column=col, value="Yes" if product.has_bim else "No"); col += 1
                    ws.cell(row=row, column=col, value="Yes" if product.has_cad else "No"); col += 1
                    ws.cell(row=row, column=col, value="Yes" if product.has_ceu else "No"); col += 1
                    ws.cell(row=row, column=col, value="Yes" if product.has_catalog else "No"); col += 1
                    ws.cell(row=row, column=col, value="Yes" if product.has_data_sheet else "No"); col += 1
                    ws.cell(row=row, column=col, value="Yes" if product.has_gallery else "No"); col += 1
                    ws.cell(row=row, column=col, value="Yes" if product.has_green else "No"); col += 1
                    ws.cell(row=row, column=col, value="Yes" if product.has_product_selector else "No"); col += 1
                    ws.cell(row=row, column=col, value="Yes" if product.has_supporting_material else "No"); col += 1

                    # Add download counts if Selenium mode was used
                    if self.use_selenium:
                        ws.cell(row=row, column=col, value=product.count_3part_spec); col += 1
                        ws.cell(row=row, column=col, value=product.count_bim); col += 1
                        ws.cell(row=row, column=col, value=product.count_cad); col += 1
                        ws.cell(row=row, column=col, value=product.count_ceu); col += 1
                        ws.cell(row=row, column=col, value=product.count_catalog); col += 1
                        ws.cell(row=row, column=col, value=product.count_gallery); col += 1
                        ws.cell(row=row, column=col, value=product.count_green); col += 1
                        ws.cell(row=row, column=col, value=product.count_other); col += 1
                        ws.cell(row=row, column=col, value=product.count_total); col += 1

                    ws.cell(row=row, column=col, value=product.url); col += 1
                    ws.cell(row=row, column=col, value="SWEETS")
                    row += 1

        # Adjust column widths
        column_widths = {
            'A': 15, 'B': 35, 'C': 15, 'D': 40, 'E': 35, 'F': 50,
            'G': 40, 'H': 20, 'I': 15, 'J': 12, 'K': 18, 'L': 18,
            'M': 30, 'N': 35, 'O': 30, 'P': 40, 'Q': 12, 'R': 10,
            'S': 10, 'T': 10, 'U': 10, 'V': 12, 'W': 10, 'X': 10,
            'Y': 15, 'Z': 18, 'AA': 60, 'AB': 10
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        try:
            wb.save(filename)
            logger.info(f"Data exported to {filename}")
        except PermissionError:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base, ext = filename.rsplit('.', 1)
            new_filename = f"{base}_{timestamp}.{ext}"
            wb.save(new_filename)
            logger.info(f"Original file was locked. Data exported to {new_filename}")
            return new_filename
        return filename

    def scrape_all(self, max_divisions: int = None, max_sections_per_division: int = None,
                   max_products_per_section: int = None, resume: bool = False):
        """Main scraping method"""
        logger.info("=" * 60)
        logger.info("STARTING SWEETS SCRAPER")
        logger.info("Source: https://sweets.construction.com/quicklinks/3partspecs")
        logger.info("=" * 60)

        # Track already scraped product URLs
        scraped_product_urls = set()

        # Check for resume
        if resume:
            checkpoint = self._load_checkpoint()
            if checkpoint:
                self._restore_from_checkpoint(checkpoint)
                for div in self.divisions:
                    for section in div.sections:
                        for product in section.products:
                            if product.phone or product.email:  # Product has been detailed
                                scraped_product_urls.add(product.url)
                logger.info(f"Resuming with {len(scraped_product_urls)} already scraped products")

        try:
            # Get all divisions
            if not self.divisions:
                self.scrape_divisions()

            divisions_to_process = self.divisions[:max_divisions] if max_divisions else self.divisions

            # First pass: count total products
            logger.info("Counting total products across all divisions...")
            total_products = 0
            for division in divisions_to_process:
                if not division.sections:
                    self.scrape_division_sections(division)

                sections_to_count = division.sections[:max_sections_per_division] if max_sections_per_division else division.sections
                for section in sections_to_count:
                    if not section.products:
                        self.scrape_section_products(section, division)

                    products = section.products[:max_products_per_section] if max_products_per_section else section.products
                    total_products += len([p for p in products if p.url not in scraped_product_urls])

            logger.info(f"Total products to scrape: {total_products:,}")
            self.progress.start(total_products)

            # Second pass: scrape product details
            for div_idx, division in enumerate(divisions_to_process):
                logger.info(f"\nProcessing division {div_idx + 1}/{len(divisions_to_process)}: {division.code} - {division.name}")

                if self.interrupted:
                    break

                sections_to_process = division.sections[:max_sections_per_division] if max_sections_per_division else division.sections

                for section in sections_to_process:
                    if self.interrupted:
                        break

                    products_to_process = section.products[:max_products_per_section] if max_products_per_section else section.products

                    for product in products_to_process:
                        if product.url in scraped_product_urls:
                            continue

                        if self.interrupted:
                            break

                        self.scrape_product_details(product)
                        self.products_scraped_count += 1
                        self.progress.update()

                        # Periodic checkpoint
                        self._maybe_save_checkpoint()

                        # Progress update every 10 products
                        if self.products_scraped_count % 10 == 0:
                            logger.info(f"  {self.progress.get_status_line()}")

                if self.interrupted:
                    break

            if not self.interrupted:
                logger.info(f"\nSWEETS scraping complete!")
                logger.info(f"Final: {self.progress.get_status_line()}")
                self._save_checkpoint()

            return self.divisions

        except Exception as e:
            logger.error(f"Error during scraping: {e}")
            self._save_checkpoint()
            raise


def main():
    """Main entry point"""
    import argparse

    parser = argparse.ArgumentParser(
        description='SWEETS Website Scraper',
        epilog='''
DOWNLOAD SECTION NOTES:
  The Downloads tab section (showing CAD, BIM, 3-Part Spec counts) is loaded
  via JavaScript/AJAX and is NOT in the initial HTML response.

  Two modes are available:
  1. DEFAULT MODE: Uses regex patterns to detect content types (Yes/No)
  2. SELENIUM MODE: Uses browser automation to capture exact download counts

  Use --use-selenium if you need exact download counts like "CAD (1)", "BIM (1)".
  This requires: pip install selenium webdriver-manager
        '''
    )
    parser.add_argument('--max-divisions', type=int, default=None,
                        help='Limit number of divisions to scrape')
    parser.add_argument('--max-sections', type=int, default=None,
                        help='Limit number of sections per division')
    parser.add_argument('--max-products', type=int, default=None,
                        help='Limit number of products per section')
    parser.add_argument('--output', type=str, default=OUTPUT_FILE,
                        help='Output Excel file path')
    parser.add_argument('--resume', action='store_true',
                        help='Resume from last checkpoint')
    parser.add_argument('--clear-checkpoint', action='store_true',
                        help='Clear existing checkpoint')
    parser.add_argument('--use-selenium', action='store_true',
                        help='Use Selenium to capture exact download counts (slower but more accurate)')

    args = parser.parse_args()

    # Initialize scraper with Selenium mode if requested
    scraper = SWEETSScraper(use_selenium=args.use_selenium)

    if args.clear_checkpoint:
        scraper.clear_checkpoint()
        print("Checkpoint cleared.")

    print("Starting SWEETS Scraper...")
    print("=" * 50)
    if args.use_selenium:
        print("Mode: SELENIUM (exact download counts)")
        print("  Captures: ALL, CATALOGS, CAD, BIM, 3-PART SPECS, GALLERIES, CEU, GREEN, OTHER")
    else:
        print("Mode: DEFAULT (pattern-based detection)")
        print("  Use --use-selenium for exact download counts")
    print("=" * 50)

    if args.resume:
        print("Resume mode enabled")

    try:
        scraper.scrape_all(
            max_divisions=args.max_divisions,
            max_sections_per_division=args.max_sections,
            max_products_per_section=args.max_products,
            resume=args.resume
        )

        # Export to Excel
        scraper.export_to_excel(args.output)

        # Clear checkpoint on success
        scraper.clear_checkpoint()

        # Clean up Selenium resources
        scraper.close()

        total_products = sum(
            len(section.products)
            for div in scraper.divisions
            for section in div.sections
        )
        print("\n" + "=" * 50)
        print(f"Scraping complete! Data saved to: {args.output}")
        print(f"Total divisions: {len(scraper.divisions)}")
        print(f"Total products: {total_products}")
        if args.use_selenium:
            print("Download counts included in Excel output.")

    except KeyboardInterrupt:
        print("\n\nInterrupted by user. Progress has been saved.")
        print("Resume with: python sweets_scraper.py --resume")
        scraper.export_to_excel(args.output)
        scraper.close()
        print(f"Partial data saved to: {args.output}")
    except Exception as e:
        print(f"\n\nError occurred: {e}")
        print("Progress saved. Resume with: python sweets_scraper.py --resume")
        try:
            scraper.export_to_excel(args.output)
            print(f"Partial data saved to: {args.output}")
        except:
            pass
        scraper.close()
        raise


if __name__ == "__main__":
    main()
