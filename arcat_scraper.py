"""
ARCAT Website Scraper
Scrapes construction specifications, companies, and contact details from arcat.com
Includes checkpoint/resume functionality for handling interruptions
"""

import requests
from bs4 import BeautifulSoup
import time
import re
import json
import os
import signal
import sys
import atexit
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from dataclasses import dataclass, field, asdict
from typing import Optional
import logging
from datetime import datetime

# Selenium imports for JavaScript-rendered content
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Constants
BASE_URL = "https://www.arcat.com"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
}
REQUEST_DELAY = 1.5  # Delay between requests to be respectful

# Checkpoint settings
CHECKPOINT_DIR = "c:/Users/user1/Desktop/PersonalGit/Scraper/checkpoints"
CHECKPOINT_INTERVAL = 10  # Save checkpoint every N companies scraped

# Retry settings for connection issues
MAX_RETRIES = 3
RETRY_DELAY_BASE = 5  # Base delay in seconds (will use exponential backoff)
REQUEST_TIMEOUT = 30  # Timeout for requests in seconds

# Partial save file
PARTIAL_SAVE_FILE = "c:/Users/user1/Desktop/PersonalGit/Scraper/ARCAT_Scraped_Data_Partial.xlsx"

# US State abbreviations to full names
STATE_ABBREV_TO_FULL = {
    'AL': 'Alabama', 'AK': 'Alaska', 'AZ': 'Arizona', 'AR': 'Arkansas',
    'CA': 'California', 'CO': 'Colorado', 'CT': 'Connecticut', 'DE': 'Delaware',
    'FL': 'Florida', 'GA': 'Georgia', 'HI': 'Hawaii', 'ID': 'Idaho',
    'IL': 'Illinois', 'IN': 'Indiana', 'IA': 'Iowa', 'KS': 'Kansas',
    'KY': 'Kentucky', 'LA': 'Louisiana', 'ME': 'Maine', 'MD': 'Maryland',
    'MA': 'Massachusetts', 'MI': 'Michigan', 'MN': 'Minnesota', 'MS': 'Mississippi',
    'MO': 'Missouri', 'MT': 'Montana', 'NE': 'Nebraska', 'NV': 'Nevada',
    'NH': 'New Hampshire', 'NJ': 'New Jersey', 'NM': 'New Mexico', 'NY': 'New York',
    'NC': 'North Carolina', 'ND': 'North Dakota', 'OH': 'Ohio', 'OK': 'Oklahoma',
    'OR': 'Oregon', 'PA': 'Pennsylvania', 'RI': 'Rhode Island', 'SC': 'South Carolina',
    'SD': 'South Dakota', 'TN': 'Tennessee', 'TX': 'Texas', 'UT': 'Utah',
    'VT': 'Vermont', 'VA': 'Virginia', 'WA': 'Washington', 'WV': 'West Virginia',
    'WI': 'Wisconsin', 'WY': 'Wyoming', 'DC': 'District of Columbia',
    # Canadian provinces
    'AB': 'Alberta', 'BC': 'British Columbia', 'MB': 'Manitoba', 'NB': 'New Brunswick',
    'NL': 'Newfoundland and Labrador', 'NS': 'Nova Scotia', 'NT': 'Northwest Territories',
    'NU': 'Nunavut', 'ON': 'Ontario', 'PE': 'Prince Edward Island', 'QC': 'Quebec',
    'SK': 'Saskatchewan', 'YT': 'Yukon'
}

# Keywords to identify associations (to be filtered out)
ASSOCIATION_KEYWORDS = [
    'association', 'institute', 'society', 'council', 'foundation',
    'federation', 'board', 'committee', 'organization', 'alliance',
    'coalition', 'consortium', 'forum', 'guild', 'league', 'union'
]


@dataclass
class Company:
    """Represents a company from ARCAT"""
    name: str
    url: str
    company_id: str = ""
    address: str = ""
    state: str = ""
    phone: str = ""
    website: str = ""
    email: str = ""
    product_expert_name: str = ""
    product_expert_phone: str = ""
    product_expert_email: str = ""
    products: list = field(default_factory=list)
    building_product_category: str = ""  # New field for building product category


@dataclass
class BuildingProductCategory:
    """Represents a building product category from ARCAT"""
    name: str
    url: str
    subcategories: list = field(default_factory=list)
    related_csi_divisions: list = field(default_factory=list)
    companies: list = field(default_factory=list)


@dataclass
class Specification:
    """Represents a specification entry"""
    csi_code: str
    title: str
    company: Company
    division: str


@dataclass
class Division:
    """Represents a CSI Division"""
    code: str
    name: str
    url: str
    specifications: list = field(default_factory=list)
    companies: list = field(default_factory=list)


class ProgressTracker:
    """Tracks scraping progress and provides ETA estimates"""

    def __init__(self):
        self.start_time = None
        self.total_companies = 0
        self.scraped_companies = 0
        self.scrape_times = []  # Track individual scrape times for better ETA
        self.last_update_time = None

    def start(self, total_companies: int = 0):
        """Start tracking progress"""
        self.start_time = time.time()
        self.total_companies = total_companies
        self.scraped_companies = 0
        self.scrape_times = []
        self.last_update_time = time.time()

    def set_total(self, total: int):
        """Update total companies count"""
        self.total_companies = total

    def update(self, count: int = 1):
        """Update progress by count"""
        current_time = time.time()
        if self.last_update_time:
            self.scrape_times.append(current_time - self.last_update_time)
            # Keep only last 100 times for moving average
            if len(self.scrape_times) > 100:
                self.scrape_times = self.scrape_times[-100:]
        self.last_update_time = current_time
        self.scraped_companies += count

    def get_percentage(self) -> float:
        """Get completion percentage"""
        if self.total_companies == 0:
            return 0.0
        return (self.scraped_companies / self.total_companies) * 100

    def get_eta_seconds(self) -> float:
        """Get estimated time remaining in seconds"""
        if not self.scrape_times or self.scraped_companies == 0:
            return 0
        avg_time = sum(self.scrape_times) / len(self.scrape_times)
        remaining = self.total_companies - self.scraped_companies
        return avg_time * remaining

    def get_eta_formatted(self) -> str:
        """Get formatted ETA string"""
        eta_seconds = self.get_eta_seconds()
        if eta_seconds <= 0:
            return "Calculating..."

        hours = int(eta_seconds // 3600)
        minutes = int((eta_seconds % 3600) // 60)
        seconds = int(eta_seconds % 60)

        if hours > 0:
            return f"{hours}h {minutes}m {seconds}s"
        elif minutes > 0:
            return f"{minutes}m {seconds}s"
        else:
            return f"{seconds}s"

    def get_elapsed_formatted(self) -> str:
        """Get formatted elapsed time"""
        if not self.start_time:
            return "0s"
        elapsed = time.time() - self.start_time
        hours = int(elapsed // 3600)
        minutes = int((elapsed % 3600) // 60)
        seconds = int(elapsed % 60)

        if hours > 0:
            return f"{hours}h {minutes}m {seconds}s"
        elif minutes > 0:
            return f"{minutes}m {seconds}s"
        else:
            return f"{seconds}s"

    def get_speed(self) -> float:
        """Get scraping speed (companies per minute)"""
        if not self.scrape_times:
            return 0.0
        avg_time = sum(self.scrape_times) / len(self.scrape_times)
        if avg_time == 0:
            return 0.0
        return 60.0 / avg_time  # companies per minute

    def get_progress_bar(self, width: int = 30) -> str:
        """Get a text-based progress bar"""
        percentage = self.get_percentage()
        filled = int(width * percentage / 100)
        bar = "█" * filled + "░" * (width - filled)
        return f"[{bar}] {percentage:.1f}%"

    def get_status_line(self) -> str:
        """Get a full status line with all progress info"""
        return (
            f"{self.get_progress_bar()} | "
            f"{self.scraped_companies:,}/{self.total_companies:,} companies | "
            f"Speed: {self.get_speed():.1f}/min | "
            f"ETA: {self.get_eta_formatted()} | "
            f"Elapsed: {self.get_elapsed_formatted()}"
        )


class ARCATScraper:
    """Main scraper class for ARCAT website"""

    def __init__(self, use_selenium: bool = False, checkpoint_file: str = None):
        self.session = requests.Session()
        self.session.headers.update(HEADERS)
        self.divisions: list[Division] = []
        self.building_product_categories: list[BuildingProductCategory] = []
        self.use_selenium = use_selenium
        self.driver = None

        # Checkpoint/resume functionality
        self.checkpoint_file = checkpoint_file or os.path.join(CHECKPOINT_DIR, "arcat_checkpoint.json")
        self.companies_scraped_count = 0
        self.interrupted = False
        self._setup_checkpoint_dir()
        self._setup_signal_handlers()

        # Progress tracking
        self.progress = ProgressTracker()
        self.output_file = None  # Will be set when scraping starts

        if use_selenium:
            self._init_selenium()

    def _setup_checkpoint_dir(self):
        """Create checkpoint directory if it doesn't exist"""
        os.makedirs(CHECKPOINT_DIR, exist_ok=True)

    def _setup_signal_handlers(self):
        """Setup handlers for graceful shutdown on interrupt"""
        def signal_handler(signum, frame):
            logger.warning(f"\nInterrupt received (signal {signum}). Saving checkpoint and partial data...")
            self.interrupted = True
            self._save_checkpoint()
            # Also save partial Excel data
            self._save_partial_excel()
            logger.info("Checkpoint and partial data saved. You can resume later with --resume flag.")
            self.close()
            sys.exit(0)

        # Handle Ctrl+C and termination signals
        signal.signal(signal.SIGINT, signal_handler)
        signal.signal(signal.SIGTERM, signal_handler)
        # Register atexit handler as backup
        atexit.register(self._emergency_save)

    def _emergency_save(self):
        """Emergency save on unexpected exit"""
        if self.companies_scraped_count > 0 and not self.interrupted:
            logger.warning("Emergency checkpoint save triggered")
            self._save_checkpoint()
            self._save_partial_excel()

    def _save_partial_excel(self):
        """Save current scraped data to partial Excel file"""
        try:
            self.export_to_excel(PARTIAL_SAVE_FILE)
            logger.info(f"Partial data saved to {PARTIAL_SAVE_FILE}")
        except Exception as e:
            logger.error(f"Failed to save partial Excel: {e}")

    def _save_checkpoint(self, mode: str = "divisions"):
        """Save current progress to checkpoint file"""
        checkpoint_data = {
            "timestamp": datetime.now().isoformat(),
            "mode": mode,
            "companies_scraped": self.companies_scraped_count,
            "divisions": [],
            "building_product_categories": [],
            "division_progress": {
                "current_division_index": 0,
                "current_company_index": 0,
                "completed_divisions": []
            },
            "category_progress": {
                "current_category_index": 0,
                "current_subcategory_index": 0,
                "current_company_index": 0,
                "completed_categories": []
            }
        }

        # Save divisions with their companies
        for div in self.divisions:
            div_data = {
                "code": div.code,
                "name": div.name,
                "url": div.url,
                "companies": []
            }
            for company in div.companies:
                div_data["companies"].append({
                    "name": company.name,
                    "url": company.url,
                    "company_id": company.company_id,
                    "address": company.address,
                    "state": company.state,
                    "phone": company.phone,
                    "website": company.website,
                    "email": company.email,
                    "product_expert_name": company.product_expert_name,
                    "product_expert_phone": company.product_expert_phone,
                    "product_expert_email": company.product_expert_email,
                    "building_product_category": company.building_product_category
                })
            checkpoint_data["divisions"].append(div_data)

        # Save building product categories with their companies
        for cat in self.building_product_categories:
            cat_data = {
                "name": cat.name,
                "url": cat.url,
                "subcategories": cat.subcategories,
                "related_csi_divisions": cat.related_csi_divisions,
                "companies": []
            }
            for company in cat.companies:
                cat_data["companies"].append({
                    "name": company.name,
                    "url": company.url,
                    "company_id": company.company_id,
                    "address": company.address,
                    "state": company.state,
                    "phone": company.phone,
                    "website": company.website,
                    "email": company.email,
                    "product_expert_name": company.product_expert_name,
                    "product_expert_phone": company.product_expert_phone,
                    "product_expert_email": company.product_expert_email,
                    "building_product_category": company.building_product_category
                })
            checkpoint_data["building_product_categories"].append(cat_data)

        # Write to file
        try:
            with open(self.checkpoint_file, 'w', encoding='utf-8') as f:
                json.dump(checkpoint_data, f, indent=2, ensure_ascii=False)
            logger.info(f"Checkpoint saved: {self.companies_scraped_count} companies to {self.checkpoint_file}")
        except Exception as e:
            logger.error(f"Failed to save checkpoint: {e}")

    def _load_checkpoint(self) -> dict:
        """Load checkpoint from file if it exists"""
        if not os.path.exists(self.checkpoint_file):
            return None

        try:
            with open(self.checkpoint_file, 'r', encoding='utf-8') as f:
                checkpoint_data = json.load(f)
            logger.info(f"Loaded checkpoint from {checkpoint_data['timestamp']}")
            logger.info(f"Previously scraped: {checkpoint_data['companies_scraped']} companies")
            return checkpoint_data
        except Exception as e:
            logger.error(f"Failed to load checkpoint: {e}")
            return None

    def _restore_from_checkpoint(self, checkpoint_data: dict):
        """Restore scraper state from checkpoint data"""
        # Restore divisions
        for div_data in checkpoint_data.get("divisions", []):
            division = Division(
                code=div_data["code"],
                name=div_data["name"],
                url=div_data["url"]
            )
            for comp_data in div_data.get("companies", []):
                company = Company(
                    name=comp_data["name"],
                    url=comp_data["url"],
                    company_id=comp_data.get("company_id", ""),
                    address=comp_data.get("address", ""),
                    state=comp_data.get("state", ""),
                    phone=comp_data.get("phone", ""),
                    website=comp_data.get("website", ""),
                    email=comp_data.get("email", ""),
                    product_expert_name=comp_data.get("product_expert_name", ""),
                    product_expert_phone=comp_data.get("product_expert_phone", ""),
                    product_expert_email=comp_data.get("product_expert_email", ""),
                    building_product_category=comp_data.get("building_product_category", "")
                )
                division.companies.append(company)
            self.divisions.append(division)

        # Restore building product categories
        for cat_data in checkpoint_data.get("building_product_categories", []):
            category = BuildingProductCategory(
                name=cat_data["name"],
                url=cat_data["url"],
                subcategories=cat_data.get("subcategories", []),
                related_csi_divisions=cat_data.get("related_csi_divisions", [])
            )
            for comp_data in cat_data.get("companies", []):
                company = Company(
                    name=comp_data["name"],
                    url=comp_data["url"],
                    company_id=comp_data.get("company_id", ""),
                    address=comp_data.get("address", ""),
                    state=comp_data.get("state", ""),
                    phone=comp_data.get("phone", ""),
                    website=comp_data.get("website", ""),
                    email=comp_data.get("email", ""),
                    product_expert_name=comp_data.get("product_expert_name", ""),
                    product_expert_phone=comp_data.get("product_expert_phone", ""),
                    product_expert_email=comp_data.get("product_expert_email", ""),
                    building_product_category=comp_data.get("building_product_category", "")
                )
                category.companies.append(company)
            self.building_product_categories.append(category)

        self.companies_scraped_count = checkpoint_data.get("companies_scraped", 0)
        logger.info(f"Restored {len(self.divisions)} divisions and {len(self.building_product_categories)} categories")

    def clear_checkpoint(self):
        """Delete checkpoint file after successful completion"""
        if os.path.exists(self.checkpoint_file):
            os.remove(self.checkpoint_file)
            logger.info("Checkpoint file cleared after successful completion")

    def _maybe_save_checkpoint(self, mode: str = "divisions"):
        """Save checkpoint if enough companies have been scraped since last save"""
        if self.companies_scraped_count > 0 and self.companies_scraped_count % CHECKPOINT_INTERVAL == 0:
            self._save_checkpoint(mode)

    def _init_selenium(self):
        """Initialize Selenium WebDriver for JavaScript-rendered content"""
        chrome_options = Options()
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        chrome_options.add_argument(f'user-agent={HEADERS["User-Agent"]}')

        logger.info("Initializing Selenium WebDriver...")
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=chrome_options)

        # Hide webdriver property to avoid detection
        self.driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
            'source': 'Object.defineProperty(navigator, "webdriver", {get: () => undefined})'
        })
        logger.info("Selenium WebDriver initialized")

    def close(self):
        """Close Selenium WebDriver"""
        if self.driver:
            self.driver.quit()
            self.driver = None
            logger.info("Selenium WebDriver closed")

    def _get_page_with_selenium(self, url: str, wait_time: int = 8) -> tuple[str, BeautifulSoup]:
        """
        Fetch a page using Selenium and wait for JavaScript to render.
        Returns (raw_html, soup) tuple.
        """
        if not self.driver:
            self._init_selenium()

        try:
            time.sleep(REQUEST_DELAY)  # Rate limiting
            self.driver.get(url)

            # Wait for the page to fully load (wait for body content)
            WebDriverWait(self.driver, wait_time).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )

            # Additional wait for dynamic content
            time.sleep(3)

            html = self.driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            return html, soup

        except Exception as e:
            logger.error(f"Selenium failed to fetch {url}: {e}")
            return "", None

    def _make_request(self, url: str) -> Optional[BeautifulSoup]:
        """Make HTTP request with error handling, rate limiting, and retry logic"""
        for attempt in range(MAX_RETRIES):
            try:
                time.sleep(REQUEST_DELAY)
                response = self.session.get(url, timeout=REQUEST_TIMEOUT)
                response.raise_for_status()
                return BeautifulSoup(response.text, 'html.parser')
            except requests.exceptions.Timeout as e:
                retry_delay = RETRY_DELAY_BASE * (2 ** attempt)  # Exponential backoff
                logger.warning(f"Timeout on {url} (attempt {attempt + 1}/{MAX_RETRIES}). Retrying in {retry_delay}s...")
                if attempt < MAX_RETRIES - 1:
                    time.sleep(retry_delay)
                else:
                    logger.error(f"Failed to fetch {url} after {MAX_RETRIES} attempts: {e}")
                    return None
            except requests.exceptions.ConnectionError as e:
                retry_delay = RETRY_DELAY_BASE * (2 ** attempt)
                logger.warning(f"Connection error on {url} (attempt {attempt + 1}/{MAX_RETRIES}). Retrying in {retry_delay}s...")
                if attempt < MAX_RETRIES - 1:
                    time.sleep(retry_delay)
                else:
                    logger.error(f"Failed to fetch {url} after {MAX_RETRIES} attempts: {e}")
                    return None
            except requests.RequestException as e:
                logger.error(f"Failed to fetch {url}: {e}")
                return None
        return None

    def scrape_divisions(self) -> list[Division]:
        """Scrape main divisions listing page"""
        url = f"{BASE_URL}/content-type/spec"
        logger.info(f"Scraping divisions from {url}")

        soup = self._make_request(url)
        if not soup:
            return []

        divisions = []

        # Find all division links - they follow pattern /content-type/spec/{name}-{code}
        division_links = soup.find_all('a', href=re.compile(r'/content-type/spec/[\w-]+-\d+$'))

        seen_urls = set()
        for link in division_links:
            href = link.get('href', '')
            if href in seen_urls:
                continue
            seen_urls.add(href)

            full_url = f"{BASE_URL}{href}" if href.startswith('/') else href
            text = link.get_text(strip=True)

            # Extract division code and name (e.g., "04 MASONRY")
            match = re.match(r'^(\d+)\s+(.+)$', text)
            if match:
                code = match.group(1)
                name = match.group(2)
                divisions.append(Division(code=code, name=name, url=full_url))
                logger.info(f"Found division: {code} - {name}")

        self.divisions = divisions
        return divisions

    def scrape_division_specs(self, division: Division) -> Division:
        """Scrape specifications from a division page"""
        logger.info(f"Scraping specs for division: {division.code} - {division.name}")

        soup = self._make_request(division.url)
        if not soup:
            return division

        # Find specification entries - look for elements with data attributes or spec listings
        # Based on the page structure, specs are in containers with company/product info

        # Method 1: Look for links containing company URLs
        company_links = soup.find_all('a', href=re.compile(r'/company/[\w-]+-\d+'))

        seen_companies = {}
        skipped_associations = 0
        for link in company_links:
            href = link.get('href', '')
            company_name = link.get_text(strip=True)

            if not company_name or href in seen_companies:
                continue

            # Filter out associations
            if self._is_association(company_name):
                skipped_associations += 1
                logger.debug(f"Skipping association: {company_name}")
                continue

            # Extract company ID from URL
            id_match = re.search(r'-(\d+)$', href)
            company_id = id_match.group(1) if id_match else ""

            full_url = f"{BASE_URL}{href}" if href.startswith('/') else href

            company = Company(
                name=company_name,
                url=full_url,
                company_id=company_id
            )
            seen_companies[href] = company

        division.companies = list(seen_companies.values())
        if skipped_associations > 0:
            logger.info(f"Filtered out {skipped_associations} associations")

        # Find spec titles/codes
        # Look for CSI code patterns like "04 01 00" or "04 21 13"
        spec_containers = soup.find_all(['div', 'section', 'article'],
                                        class_=re.compile(r'spec|product|listing', re.I))

        # Also search in text for CSI codes
        all_text = soup.get_text()
        csi_codes = re.findall(r'\b(\d{2}\s+\d{2}\s+\d{2})\b', all_text)

        for code in set(csi_codes):
            # Find associated title near this code
            code_pattern = re.escape(code)
            title_match = re.search(rf'{code_pattern}\s*[-–]\s*([^\n]+)', all_text)
            title = title_match.group(1).strip() if title_match else ""

            if title:
                spec = Specification(
                    csi_code=code,
                    title=title,
                    company=None,  # Will be linked later
                    division=division.name
                )
                division.specifications.append(spec)

        logger.info(f"Found {len(division.companies)} companies and {len(division.specifications)} specs")
        return division

    def _extract_state_from_address(self, address: str) -> str:
        """Extract state abbreviation from address and convert to full name."""
        if not address:
            return ""

        # Look for state abbreviation pattern (2 letters before zip code)
        state_match = re.search(r',\s*([A-Z]{2})\s*\d{5}', address)
        if state_match:
            abbrev = state_match.group(1).upper()
            return STATE_ABBREV_TO_FULL.get(abbrev, abbrev)

        return ""

    def _is_association(self, company_name: str) -> bool:
        """Check if company name indicates it's an association"""
        name_lower = company_name.lower()
        return any(keyword in name_lower for keyword in ASSOCIATION_KEYWORDS)

    def _extract_from_nuxt_data(self, html: str) -> dict:
        """Extract company data from embedded NUXT/JavaScript serialized data.

        The ARCAT site embeds company data in a serialized format like:
        ,"Goldilox USA",null,1,"13 Pfeiffer Rd."," Boerne","TX","78006","830-981-2210","info@goldiloxusa.com",...,"https://www.goldiloxusa.com"
        """
        data = {
            'address': '',
            'city': '',
            'state': '',
            'zip': '',
            'phone': '',
            'email': '',
            'website': ''
        }

        # Pattern to match the company data sequence in the serialized NUXT data
        # Format: "address","city","state","zip",[null|phone],"phone","email"
        # Some entries have null between ZIP and phone, others don't

        # Pattern 1: With null between ZIP and phone
        # "245 Beback Inn Rd.","San Marcos","TX","78666",null,"704-408-6211","email@domain.com"
        pattern_with_null = r'"(\d+[^"]*(?:Rd|St|Ave|Dr|Blvd|Way|Ln|Pkwy|Place|Pl|Circle|Ct|Court|Road|Street|Avenue|Drive|Boulevard|Lane|Parkway)\.?[^"]*)",\s*"([^"]*)",\s*"([A-Z]{2})",\s*"(\d{5}(?:-\d{4})?)",\s*null,\s*"([^"]*)",\s*"([^"]*@[^"]*)"'

        match = re.search(pattern_with_null, html, re.IGNORECASE)
        if match:
            data['address'] = match.group(1).strip()
            data['city'] = match.group(2).strip()
            data['state'] = match.group(3).strip()
            data['zip'] = match.group(4).strip()
            data['phone'] = match.group(5).strip()
            data['email'] = match.group(6).strip()

        # Pattern 2: Without null (phone directly after ZIP)
        # "13 Pfeiffer Rd.","Boerne","TX","78006","830-981-2210","email@domain.com"
        if not data['address']:
            pattern_direct = r'"(\d+[^"]*(?:Rd|St|Ave|Dr|Blvd|Way|Ln|Pkwy|Place|Pl|Circle|Ct|Court|Road|Street|Avenue|Drive|Boulevard|Lane|Parkway)\.?[^"]*)",\s*"([^"]*)",\s*"([A-Z]{2})",\s*"(\d{5}(?:-\d{4})?)",\s*"([^"]*)",\s*"([^"]*@[^"]*)"'

            match = re.search(pattern_direct, html, re.IGNORECASE)
            if match:
                data['address'] = match.group(1).strip()
                data['city'] = match.group(2).strip()
                data['state'] = match.group(3).strip()
                data['zip'] = match.group(4).strip()
                data['phone'] = match.group(5).strip()
                data['email'] = match.group(6).strip()

        # Pattern 3: Simpler pattern for address without requiring email
        if not data['address']:
            # Try: "address","city","ST","ZIP","phone" or "address","city","ST","ZIP",null,"phone"
            pattern3 = r'"(\d+[^"]*(?:Rd|St|Ave|Dr|Blvd|Way|Ln|Pkwy|Place|Pl|Circle|Ct|Court|Road|Street|Avenue|Drive|Boulevard|Lane|Parkway)\.?[^"]*)",\s*"([^"]*)",\s*"([A-Z]{2})",\s*"(\d{5}(?:-\d{4})?)",\s*(?:null,\s*)?"(\d{3}[-.]?\d{3}[-.]?\d{4})"'
            match3 = re.search(pattern3, html, re.IGNORECASE)
            if match3:
                data['address'] = match3.group(1).strip()
                data['city'] = match3.group(2).strip()
                data['state'] = match3.group(3).strip()
                data['zip'] = match3.group(4).strip()
                data['phone'] = match3.group(5).strip()

        # Pattern 4: Canadian addresses with "BC, Canada" format and Canadian postal code
        # "13731 Mayfield Pl.","Richmond","BC, Canada","V6V 2G9","800-961-4477","604-273-5265","email"
        if not data['address']:
            pattern_canada = r'"(\d+[^"]*(?:Rd|St|Ave|Dr|Blvd|Way|Ln|Pkwy|Place|Pl|Circle|Ct|Court|Road|Street|Avenue|Drive|Boulevard|Lane|Parkway)\.?[^"]*)",\s*"([^"]*)",\s*"([A-Z]{2}),?\s*Canada",\s*"([A-Z]\d[A-Z]\s*\d[A-Z]\d)",\s*"([^"]*)",\s*"([^"]*)",\s*"([^"]*@[^"]*)"'
            match_canada = re.search(pattern_canada, html, re.IGNORECASE)
            if match_canada:
                data['address'] = match_canada.group(1).strip()
                data['city'] = match_canada.group(2).strip()
                data['state'] = match_canada.group(3).strip()  # Will be "BC" etc
                data['zip'] = match_canada.group(4).strip()
                # For Canadian, first phone is toll-free, second is local - use toll-free
                data['phone'] = match_canada.group(5).strip() or match_canada.group(6).strip()
                data['email'] = match_canada.group(7).strip()

        # Extract website URL separately - look for https:// URLs that aren't arcat
        website_pattern = r'"(https?://(?:www\.)?(?!.*arcat\.com)[a-zA-Z0-9][-a-zA-Z0-9]*\.[a-zA-Z]{2,}(?:/[^"]*)?)"'
        website_matches = re.findall(website_pattern, html)
        for url in website_matches:
            # Skip social media, tracking, schema.org, and other non-company URLs
            skip_domains = ['facebook', 'twitter', 'linkedin', 'google', 'youtube', 'clarity',
                           'analytics', 'schema.org', 'w3.org', 'cloudflare', 'jsdelivr',
                           'googleapis', 'gstatic', 'microsoft', 'bing', 'doubleclick']
            if not any(x in url.lower() for x in skip_domains):
                data['website'] = url
                break

        return data

    def _extract_from_rendered_html(self, soup: BeautifulSoup) -> dict:
        """Extract company data from fully rendered HTML (after JavaScript execution)"""
        data = {
            'address': '',
            'city': '',
            'state': '',
            'phone': '',
            'email': '',
            'website': '',
            'product_expert_name': '',
            'product_expert_phone': '',
            'product_expert_email': ''
        }

        # Look for contact info in the rendered page
        # The company page typically has address, phone, email in visible text

        page_text = soup.get_text(separator=' ', strip=True)

        # Try to find address patterns
        address_patterns = [
            # Full address with state and zip
            r'(\d+[\w\s\.\,]+(?:Rd|Road|St|Street|Ave|Avenue|Dr|Drive|Blvd|Boulevard|Way|Ln|Lane|Pkwy|Parkway|Place|Pl|Circle|Court|Ct)\.?[\w\s\.\,]+,?\s*[A-Z]{2}\s+\d{5}(?:-\d{4})?)',
            # PO Box
            r'(P\.?\s*O\.?\s*Box\s+\d+[\w\s\.\,]+,?\s*[A-Z]{2}\s+\d{5})',
        ]

        for pattern in address_patterns:
            match = re.search(pattern, page_text, re.IGNORECASE)
            if match:
                addr = match.group(1).strip()
                data['address'] = re.sub(r'\s+', ' ', addr)  # Normalize whitespace

                # Extract state
                state_match = re.search(r',?\s*([A-Z]{2})\s+\d{5}', addr)
                if state_match:
                    data['state'] = state_match.group(1)
                break

        # Phone patterns
        phone_patterns = [
            r'(?:Phone|Tel|Telephone)[:\s]*(\d{3}[-.\s]?\d{3}[-.\s]?\d{4})',
            r'(?:Toll\s*Free)[:\s]*(\d{3}[-.\s]?\d{3}[-.\s]?\d{4})',
            r'(\d{3}[-.\s]\d{3}[-.\s]\d{4})',
        ]

        for pattern in phone_patterns:
            match = re.search(pattern, page_text, re.IGNORECASE)
            if match:
                data['phone'] = match.group(1).strip()
                break

        # Email - look for mailto links first, then patterns
        mailto_link = soup.find('a', href=re.compile(r'^mailto:'))
        if mailto_link:
            href = mailto_link.get('href', '')
            email_match = re.search(r'mailto:([^\?]+)', href)
            if email_match:
                data['email'] = email_match.group(1).strip()
        else:
            email_match = re.search(r'\b([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})\b', page_text)
            if email_match:
                email = email_match.group(1)
                if not any(x in email.lower() for x in ['noreply', 'donotreply', 'unsubscribe', '@arcat']):
                    data['email'] = email

        # Website - look for external links
        excluded_domains = ['facebook.com', 'twitter.com', 'linkedin.com', 'instagram.com',
                           'youtube.com', 'arcat.com', 'pinterest.com', 'tiktok.com',
                           'google.com', 'clarity.ms']
        for link in soup.find_all('a', href=re.compile(r'^https?://')):
            href = link.get('href', '')
            if any(domain in href.lower() for domain in excluded_domains):
                continue
            if '/spec-doc/' in href or '/specification/' in href:
                continue
            # Make sure it looks like a company website
            if 'www.' in href or re.search(r'\.[a-z]{2,4}/?$', href):
                data['website'] = href
                break

        return data

    def scrape_company_details(self, company: Company) -> Company:
        """Scrape detailed company information using Selenium for dynamic content"""
        logger.info(f"Scraping company: {company.name}")

        html = ""
        soup = None

        # Use Selenium if enabled (for JavaScript-rendered content)
        if self.use_selenium and self.driver:
            html, soup = self._get_page_with_selenium(company.url)
            if soup:
                # Extract from rendered HTML
                rendered_data = self._extract_from_rendered_html(soup)
                if rendered_data['address']:
                    company.address = rendered_data['address']
                    # Convert state to full name
                    if rendered_data['state']:
                        company.state = STATE_ABBREV_TO_FULL.get(rendered_data['state'].upper(), rendered_data['state'])
                if rendered_data['phone']:
                    company.phone = rendered_data['phone']
                if rendered_data['email']:
                    company.email = rendered_data['email']
                if rendered_data['website']:
                    company.website = rendered_data['website']

        # Fallback to requests if Selenium didn't get the data or isn't enabled
        if not company.address:
            try:
                time.sleep(REQUEST_DELAY)
                response = self.session.get(company.url, timeout=30)
                response.raise_for_status()
                html = response.text
                soup = BeautifulSoup(html, 'html.parser')
            except requests.RequestException as e:
                logger.error(f"Failed to fetch {company.url}: {e}")
                return company

        # First, try to extract from embedded NUXT/JavaScript data
        nuxt_data = self._extract_from_nuxt_data(html)

        if nuxt_data['address']:
            # Build full address from components
            city = nuxt_data['city'].strip()
            state = nuxt_data['state'].strip()
            zip_code = nuxt_data['zip'].strip()
            company.address = f"{nuxt_data['address']}, {city}, {state} {zip_code}".strip(', ')

            # Convert state abbreviation to full name
            company.state = STATE_ABBREV_TO_FULL.get(state.upper(), state)

        if nuxt_data['phone']:
            company.phone = nuxt_data['phone']

        if nuxt_data['email']:
            company.email = nuxt_data['email']

        if nuxt_data['website']:
            company.website = nuxt_data['website']

        # If NUXT extraction failed, fall back to traditional HTML parsing
        if not company.address:
            page_text = soup.get_text()

            # Address patterns - more specific
            address_patterns = [
                # Full address with zip
                r'(\d+\s+[\w\s\.]+(?:Ave|Avenue|St|Street|Rd|Road|Dr|Drive|Blvd|Boulevard|Pkwy|Parkway|Way|Ln|Lane|Circle|Court|Ct|Place|Pl)\.?[,\s]+[\w\s]+,\s*[A-Z]{2}\s*\d{5}(?:-\d{4})?)',
                # PO Box
                r'(P\.?\s*O\.?\s*Box\s+\d+[,\s]+[\w\s]+,\s*[A-Z]{2}\s*\d{5})',
            ]

            for pattern in address_patterns:
                addr_match = re.search(pattern, page_text, re.IGNORECASE)
                if addr_match:
                    company.address = addr_match.group(1).strip()
                    break

            # Extract state from address and convert to full name
            if company.address:
                company.state = self._extract_state_from_address(company.address)

        # Fallback for phone if not extracted from NUXT
        if not company.phone:
            page_text = soup.get_text()
            phone_patterns = [
                r'(?:Phone|Tel|Telephone)[:\s]*(\d{3}[-.\s]?\d{3}[-.\s]?\d{4})',
                r'(?:Toll\s*Free)[:\s]*(\d{3}[-.\s]?\d{3}[-.\s]?\d{4})',
                r'(\(\d{3}\)\s*\d{3}[-.\s]?\d{4})',
                r'(?<!\d)(\d{3}[-.\s]\d{3}[-.\s]\d{4})(?!\d)',
            ]

            for pattern in phone_patterns:
                phone_match = re.search(pattern, page_text, re.IGNORECASE)
                if phone_match:
                    company.phone = phone_match.group(1).strip()
                    break

        # Fallback for website if not extracted from NUXT
        if not company.website:
            excluded_domains = ['facebook.com', 'twitter.com', 'linkedin.com', 'instagram.com',
                               'youtube.com', 'arcat.com', 'pinterest.com', 'tiktok.com']

            website_links = soup.find_all('a', href=re.compile(r'^https?://'))
            for link in website_links:
                href = link.get('href', '')
                if any(domain in href.lower() for domain in excluded_domains):
                    continue
                if '/spec-doc/' in href or '/specification/' in href:
                    continue
                if 'www.' in href or any(ext in href for ext in ['.com', '.org', '.net', '.us', '.co']):
                    company.website = href
                    break

        # Fallback for email if not extracted from NUXT - be more careful with pattern
        if not company.email:
            # Look for mailto: links first (most reliable)
            mailto_match = re.search(r'href="mailto:([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})"', html)
            if mailto_match:
                company.email = mailto_match.group(1).strip()
            else:
                # Look for info@ or contact@ style emails in the main content
                page_text = soup.get_text()
                email_match = re.search(r'\b([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})\b', page_text)
                if email_match:
                    email = email_match.group(1)
                    if not any(x in email.lower() for x in ['noreply', 'donotreply', 'unsubscribe']):
                        company.email = email

        # Product Expert - multiple extraction methods
        # Method 1: Look for HTML pattern like: <p>Name<br>Phone<br><a href="mailto:email">
        expert_html_pattern = r'<p>([A-Z][a-z]+\s+[A-Z][a-z]+)<br[^>]*>(\d{3}[-.\s]?\d{3}[-.\s]?\d{4})<br[^>]*><a[^>]*href="mailto:([^"]+)"'
        expert_html_match = re.search(expert_html_pattern, html)
        if expert_html_match:
            company.product_expert_name = expert_html_match.group(1).strip()
            company.product_expert_phone = expert_html_match.group(2).strip()
            company.product_expert_email = expert_html_match.group(3).strip()
        else:
            # Method 2: Look for NUXT data pattern
            expert_pattern = r'Product\s*Expert[^"]*"([A-Z][a-z]+\s+[A-Z][a-z]+)"[^"]*"(\d{3}[-.\s]?\d{3}[-.\s]?\d{4})"[^"]*"([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})"'
            expert_match = re.search(expert_pattern, html, re.IGNORECASE)
            if expert_match:
                company.product_expert_name = expert_match.group(1).strip()
                company.product_expert_phone = expert_match.group(2).strip()
                company.product_expert_email = expert_match.group(3).strip()
            else:
                # Method 3: Look in serialized JSON for name,repId,phone,email pattern
                alt_expert_pattern = r'"([A-Z][a-z]+\s+[A-Z][a-z]+)",\d+,"(\d{3}[-.\s]?\d{3}[-.\s]?\d{4})","([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})"'
                alt_match = re.search(alt_expert_pattern, html)
                if alt_match:
                    company.product_expert_name = alt_match.group(1).strip()
                    company.product_expert_phone = alt_match.group(2).strip()
                    company.product_expert_email = alt_match.group(3).strip()

        logger.info(f"Scraped: {company.name} | Address: {company.address[:30] if company.address else 'N/A'}... | State: {company.state} | Phone: {company.phone}")
        return company

    def scrape_building_product_categories(self) -> list[BuildingProductCategory]:
        """Scrape all building product categories from the main categories page"""
        url = f"{BASE_URL}/products/building_products_categories"
        logger.info(f"Scraping building product categories from {url}")

        soup = self._make_request(url)
        if not soup:
            return []

        categories = []

        # Find all category links - they follow pattern /products/{category_name}
        category_links = soup.find_all('a', href=re.compile(r'^/products/[a-z_]+$'))

        seen_urls = set()
        for link in category_links:
            href = link.get('href', '')
            if href in seen_urls or href == '/products/building_products_categories':
                continue
            seen_urls.add(href)

            full_url = f"{BASE_URL}{href}"
            # Extract category name from URL and format it nicely
            category_name = href.split('/')[-1].replace('_', ' ').title()

            category = BuildingProductCategory(
                name=category_name,
                url=full_url
            )
            categories.append(category)
            logger.info(f"Found category: {category_name}")

        self.building_product_categories = categories
        return categories

    def scrape_related_csi_divisions(self) -> list[Division]:
        """
        Scrape Related CSI Divisions from the Building Product Categories page.

        Source: https://www.arcat.com/products/building_products_categories
        These are the divisions shown in the "Related CSI Divisions" sidebar.

        Returns divisions with URLs like:
        /content-type/product/existing-conditions-02/existing-conditions-020000
        """
        url = f"{BASE_URL}/products/building_products_categories"
        logger.info(f"Scraping Related CSI Divisions from {url}")

        soup = self._make_request(url)
        if not soup:
            return []

        divisions = []

        # Find CSI division links matching pattern: /content-type/product/xxx-XX/xxx-XXXXXX
        csi_links = soup.find_all('a', href=re.compile(r'/content-type/product/[\w-]+-\d+/[\w-]+-\d+'))

        seen_urls = set()
        for link in csi_links:
            href = link.get('href', '')
            if href in seen_urls:
                continue
            seen_urls.add(href)

            text = link.get_text(strip=True)
            full_url = f"{BASE_URL}{href}"

            # Parse division code and name from text like "02 - EXISTING CONDITIONS"
            match = re.match(r'^(\d+)\s*-\s*(.+)$', text)
            if match:
                code = match.group(1)
                name = match.group(2).strip()

                # Format code as "XX 00 00" (e.g., "02 00 00")
                formatted_code = f"{code.zfill(2)} 00 00"

                division = Division(
                    code=formatted_code,
                    name=name,
                    url=full_url
                )
                divisions.append(division)
                logger.info(f"Found division: {formatted_code} - {name}")

        self.divisions = divisions
        return divisions

    def scrape_division_manufacturers(self, division: Division) -> Division:
        """
        Scrape manufacturers/companies from a CSI division page.

        Source URLs like:
        https://www.arcat.com/content-type/product/existing-conditions-02/existing-conditions-020000

        This gets the list of companies (manufacturers) for a specific CSI division.
        """
        logger.info(f"Scraping manufacturers for division: {division.code} - {division.name}")

        soup = self._make_request(division.url)
        if not soup:
            return division

        # Find company links - pattern: /company/{company-slug}-{id}
        # Filter to only get main company links, not sub-pages like /cad, /bim, etc.
        company_links = soup.find_all('a', href=re.compile(r'/company/[\w-]+-\d+(?:\?|$)'))

        seen_companies = {}
        skipped_associations = 0

        for link in company_links:
            href = link.get('href', '')
            company_name = link.get_text(strip=True)

            # Skip empty names or sub-page links
            if not company_name or '/cad' in href or '/bim' in href or '/spec' in href:
                continue

            # Clean up URL (remove query params for deduplication)
            base_href = href.split('?')[0]
            if base_href in seen_companies:
                continue

            # Filter out associations
            if self._is_association(company_name):
                skipped_associations += 1
                logger.debug(f"Skipping association: {company_name}")
                continue

            # Extract company ID from URL
            id_match = re.search(r'-(\d+)(?:\?|$)', href)
            company_id = id_match.group(1) if id_match else ""

            full_url = f"{BASE_URL}{href}" if href.startswith('/') else href

            company = Company(
                name=company_name,
                url=full_url,
                company_id=company_id,
                building_product_category=f"{division.code} - {division.name}"
            )
            seen_companies[base_href] = company

        division.companies = list(seen_companies.values())

        if skipped_associations > 0:
            logger.info(f"Filtered out {skipped_associations} associations")

        logger.info(f"Found {len(division.companies)} companies in division {division.code}")
        return division

    def scrape_category_subcategories(self, category: BuildingProductCategory) -> BuildingProductCategory:
        """Scrape subcategories and related CSI divisions for a building product category"""
        logger.info(f"Scraping subcategories for category: {category.name}")

        soup = self._make_request(category.url)
        if not soup:
            return category

        # Find subcategory links - they lead to /manufacturers/{subcategory}
        subcategory_links = soup.find_all('a', href=re.compile(r'^/manufacturers/[a-z_-]+$'))

        seen_urls = set()
        for link in subcategory_links:
            href = link.get('href', '')
            if href in seen_urls:
                continue
            seen_urls.add(href)

            full_url = f"{BASE_URL}{href}"
            subcategory_name = href.split('/')[-1].replace('_', ' ').replace('-', ' ').title()
            category.subcategories.append({
                'name': subcategory_name,
                'url': full_url
            })

        # Find Related CSI Divisions
        csi_links = soup.find_all('a', href=re.compile(r'/content-type/spec/[\w-]+-\d+$'))
        seen_csi = set()
        for link in csi_links:
            href = link.get('href', '')
            if href in seen_csi:
                continue
            seen_csi.add(href)

            text = link.get_text(strip=True)
            full_url = f"{BASE_URL}{href}"
            category.related_csi_divisions.append({
                'name': text,
                'url': full_url
            })

        logger.info(f"Found {len(category.subcategories)} subcategories and {len(category.related_csi_divisions)} CSI divisions for {category.name}")
        return category

    def scrape_manufacturers_page(self, subcategory_url: str, category_name: str) -> list[Company]:
        """Scrape companies from a manufacturers/subcategory page"""
        logger.info(f"Scraping manufacturers from {subcategory_url}")

        soup = self._make_request(subcategory_url)
        if not soup:
            return []

        companies = []

        # Find company links - pattern: /company/{company-slug}-{id}
        company_links = soup.find_all('a', href=re.compile(r'/company/[\w-]+-\d+$'))

        seen_companies = {}
        skipped_associations = 0

        for link in company_links:
            href = link.get('href', '')
            company_name = link.get_text(strip=True)

            if not company_name or href in seen_companies:
                continue

            # Filter out associations
            if self._is_association(company_name):
                skipped_associations += 1
                logger.debug(f"Skipping association: {company_name}")
                continue

            # Extract company ID from URL
            id_match = re.search(r'-(\d+)$', href)
            company_id = id_match.group(1) if id_match else ""

            full_url = f"{BASE_URL}{href}" if href.startswith('/') else href

            company = Company(
                name=company_name,
                url=full_url,
                company_id=company_id,
                building_product_category=category_name  # Set the category
            )
            seen_companies[href] = company

        companies = list(seen_companies.values())

        if skipped_associations > 0:
            logger.info(f"Filtered out {skipped_associations} associations")

        logger.info(f"Found {len(companies)} companies in {subcategory_url}")
        return companies

    def scrape_building_products_all(self, max_categories: int = None, max_subcategories_per_category: int = None,
                                      max_companies_per_subcategory: int = None, resume: bool = False):
        """
        Scrape all companies through building product categories navigation with checkpoint/resume support

        Args:
            max_categories: Limit number of categories to scrape (for testing)
            max_subcategories_per_category: Limit subcategories per category (for testing)
            max_companies_per_subcategory: Limit companies per subcategory (for testing)
            resume: If True, try to resume from last checkpoint
        """
        logger.info("Starting Building Product Categories scrape")

        # Track which companies we've already scraped (for resume)
        scraped_company_urls = set()

        # Check for resume
        if resume:
            checkpoint = self._load_checkpoint()
            if checkpoint:
                self._restore_from_checkpoint(checkpoint)
                # Build set of already scraped company URLs from categories
                for cat in self.building_product_categories:
                    for company in cat.companies:
                        if company.address:  # Company has been fully scraped
                            scraped_company_urls.add(company.url)
                logger.info(f"Resuming with {len(scraped_company_urls)} already scraped companies")

        try:
            # Get all building product categories (only if not resuming or no categories loaded)
            if not self.building_product_categories:
                self.scrape_building_product_categories()

            categories_to_process = self.building_product_categories[:max_categories] if max_categories else self.building_product_categories

            # First pass: count total companies to scrape for progress tracking
            logger.info("Counting total companies to scrape...")
            total_companies = 0
            for category in categories_to_process:
                if not category.subcategories:
                    self.scrape_category_subcategories(category)
                subcats = category.subcategories[:max_subcategories_per_category] if max_subcategories_per_category else category.subcategories
                for subcat in subcats:
                    companies = self.scrape_manufacturers_page(subcat['url'], category.name)
                    count = len(companies[:max_companies_per_subcategory] if max_companies_per_subcategory else companies)
                    total_companies += count

            # Subtract already scraped companies
            total_companies -= len(scraped_company_urls)
            if total_companies < 0:
                total_companies = 0

            logger.info(f"Total companies to scrape: {total_companies:,}")
            self.progress.start(total_companies)

            all_companies = []  # Track all companies with their categories

            for cat_idx, category in enumerate(categories_to_process):
                logger.info(f"\nProcessing category {cat_idx + 1}/{len(categories_to_process)}: {category.name}")

                # Check if interrupted
                if self.interrupted:
                    break

                # Subcategories already populated in counting pass
                subcategories_to_process = category.subcategories[:max_subcategories_per_category] if max_subcategories_per_category else category.subcategories

                for sub_idx, subcategory in enumerate(subcategories_to_process):
                    logger.info(f"  Subcategory {sub_idx + 1}/{len(subcategories_to_process)}: {subcategory['name']}")

                    if self.interrupted:
                        break

                    # Get companies from this subcategory's manufacturers page
                    companies = self.scrape_manufacturers_page(subcategory['url'], category.name)

                    companies_to_process = companies[:max_companies_per_subcategory] if max_companies_per_subcategory else companies

                    # Get details for each company
                    for company in companies_to_process:
                        # Skip if already scraped (resume mode)
                        if company.url in scraped_company_urls:
                            logger.debug(f"Skipping already scraped: {company.name}")
                            continue

                        if self.interrupted:
                            break

                        self.scrape_company_details(company)
                        all_companies.append(company)
                        self.companies_scraped_count += 1
                        self.progress.update()

                        # Periodic checkpoint save
                        self._maybe_save_checkpoint("categories")

                        # Progress update every 10 companies
                        if self.companies_scraped_count % 10 == 0:
                            logger.info(f"  {self.progress.get_status_line()}")

                    # Add newly scraped companies to category
                    new_companies = [c for c in companies_to_process if c.url not in scraped_company_urls]
                    category.companies.extend(new_companies)

                if self.interrupted:
                    break

            if not self.interrupted:
                logger.info(f"\nBuilding Products scraping complete! Found {len(all_companies)} companies")
                logger.info(f"Final: {self.progress.get_status_line()}")
                # Save final checkpoint
                self._save_checkpoint("categories")

            return all_companies

        except Exception as e:
            logger.error(f"Error during scraping: {e}")
            # Save checkpoint on error
            self._save_checkpoint("categories")
            raise
        finally:
            # Always close the browser when done
            self.close()

    def scrape_all(self, max_divisions: int = None, max_companies_per_division: int = None, resume: bool = False):
        """
        Main method to scrape all data with checkpoint/resume support

        Args:
            max_divisions: Limit number of divisions to scrape (for testing)
            max_companies_per_division: Limit companies per division (for testing)
            resume: If True, try to resume from last checkpoint
        """
        logger.info("Starting full ARCAT scrape")

        # Track which companies we've already scraped (for resume)
        scraped_company_urls = set()

        # Check for resume
        if resume:
            checkpoint = self._load_checkpoint()
            if checkpoint:
                self._restore_from_checkpoint(checkpoint)
                # Build set of already scraped company URLs
                for div in self.divisions:
                    for company in div.companies:
                        if company.address:  # Company has been fully scraped
                            scraped_company_urls.add(company.url)
                logger.info(f"Resuming with {len(scraped_company_urls)} already scraped companies")

        try:
            # Get all divisions (only if not resuming or no divisions loaded)
            if not self.divisions:
                self.scrape_divisions()

            divisions_to_process = self.divisions[:max_divisions] if max_divisions else self.divisions

            for div_idx, division in enumerate(divisions_to_process):
                logger.info(f"Processing division {div_idx + 1}/{len(divisions_to_process)}: {division.code} - {division.name}")

                # Get specs and companies for each division (only if not already populated)
                if not division.companies:
                    self.scrape_division_specs(division)

                # Get details for each company
                companies_to_process = division.companies[:max_companies_per_division] if max_companies_per_division else division.companies

                for comp_idx, company in enumerate(companies_to_process):
                    # Skip if already scraped (resume mode)
                    if company.url in scraped_company_urls:
                        logger.debug(f"Skipping already scraped: {company.name}")
                        continue

                    # Check if interrupted
                    if self.interrupted:
                        break

                    self.scrape_company_details(company)
                    self.companies_scraped_count += 1

                    # Periodic checkpoint save
                    self._maybe_save_checkpoint("divisions")

                    # Progress update every 25 companies
                    if self.companies_scraped_count % 25 == 0:
                        logger.info(f"Progress: {self.companies_scraped_count} companies scraped")

                if self.interrupted:
                    break

            if not self.interrupted:
                logger.info("Division scraping complete!")
                # Save final checkpoint
                self._save_checkpoint("divisions")
        except Exception as e:
            logger.error(f"Error during scraping: {e}")
            # Save checkpoint on error
            self._save_checkpoint("divisions")
            raise
        finally:
            # Always close the browser when done
            self.close()

        return self.divisions

    def export_to_excel(self, filename: str = "ARCAT_Data.xlsx"):
        """Export scraped data to Excel file - one company per row"""
        logger.info(f"Exporting data to {filename}")

        wb = Workbook()
        ws = wb.active
        ws.title = "ARCAT Data"

        # Define headers - each field in its own column
        headers = [
            "Division Code",
            "Division Name",
            "Building Product Categories",  # New column
            "Company",
            "Address",
            "State",
            "Phone",
            "Website",
            "Email",
            "Product Expert Email",
            "Product Expert Phone",
            "Company URL",
            "Source"
        ]

        # Style for headers
        header_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.border = thin_border

        # Write data - one company per row
        row = 2
        for division in self.divisions:
            # Division code may already be formatted as "XX 00 00" or just "XX"
            # If it's just digits, format it; otherwise use as-is
            if division.code.isdigit():
                formatted_code = f"{division.code.zfill(2)} 00 00"
            else:
                formatted_code = division.code  # Already formatted or special format

            if not division.companies:
                # Write division even if no companies found
                ws.cell(row=row, column=1, value=formatted_code)
                ws.cell(row=row, column=2, value=division.name)
                ws.cell(row=row, column=13, value="ARCAT")
                row += 1
                continue

            # Each company gets its own row
            for company in division.companies:
                ws.cell(row=row, column=1, value=formatted_code)
                ws.cell(row=row, column=2, value=division.name)
                ws.cell(row=row, column=3, value=company.building_product_category if company.building_product_category else "")
                ws.cell(row=row, column=4, value=company.name)
                ws.cell(row=row, column=5, value=company.address if company.address else "")
                ws.cell(row=row, column=6, value=company.state if company.state else "")
                ws.cell(row=row, column=7, value=company.phone if company.phone else "")
                ws.cell(row=row, column=8, value=company.website if company.website else "")
                ws.cell(row=row, column=9, value=company.email if company.email else "")
                ws.cell(row=row, column=10, value=company.product_expert_email if company.product_expert_email else "")
                ws.cell(row=row, column=11, value=company.product_expert_phone if company.product_expert_phone else "")
                ws.cell(row=row, column=12, value=company.url if company.url else "")
                ws.cell(row=row, column=13, value="ARCAT")
                row += 1

        # Also export building product categories data if available
        for category in self.building_product_categories:
            for company in category.companies:
                ws.cell(row=row, column=1, value="")  # No division code for category-based entries
                ws.cell(row=row, column=2, value="")  # No division name for category-based entries
                ws.cell(row=row, column=3, value=company.building_product_category)
                ws.cell(row=row, column=4, value=company.name)
                ws.cell(row=row, column=5, value=company.address if company.address else "")
                ws.cell(row=row, column=6, value=company.state if company.state else "")
                ws.cell(row=row, column=7, value=company.phone if company.phone else "")
                ws.cell(row=row, column=8, value=company.website if company.website else "")
                ws.cell(row=row, column=9, value=company.email if company.email else "")
                ws.cell(row=row, column=10, value=company.product_expert_email if company.product_expert_email else "")
                ws.cell(row=row, column=11, value=company.product_expert_phone if company.product_expert_phone else "")
                ws.cell(row=row, column=12, value=company.url if company.url else "")
                ws.cell(row=row, column=13, value="ARCAT")
                row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 15  # Division Code
        ws.column_dimensions['B'].width = 40  # Division Name
        ws.column_dimensions['C'].width = 30  # Building Product Categories
        ws.column_dimensions['D'].width = 50  # Company
        ws.column_dimensions['E'].width = 50  # Address
        ws.column_dimensions['F'].width = 20  # State
        ws.column_dimensions['G'].width = 18  # Phone
        ws.column_dimensions['H'].width = 35  # Website
        ws.column_dimensions['I'].width = 30  # Email
        ws.column_dimensions['J'].width = 30  # Product Expert Email
        ws.column_dimensions['K'].width = 18  # Product Expert Phone
        ws.column_dimensions['L'].width = 50  # Company URL
        ws.column_dimensions['M'].width = 10  # Source

        try:
            wb.save(filename)
            logger.info(f"Data exported to {filename}")
        except PermissionError:
            # File might be open, try with timestamp
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base, ext = filename.rsplit('.', 1)
            new_filename = f"{base}_{timestamp}.{ext}"
            wb.save(new_filename)
            logger.info(f"Original file was locked. Data exported to {new_filename}")
            return new_filename
        return filename


def main():
    """Main entry point"""
    import argparse

    parser = argparse.ArgumentParser(description='ARCAT Website Scraper with checkpoint/resume support')
    parser.add_argument('--mode', choices=['divisions', 'categories', 'both', 'csi-only'], default='csi-only',
                        help='Scraping mode: csi-only (recommended - ~28 CSI divisions), divisions, categories, or both')
    parser.add_argument('--max-divisions', type=int, default=None,
                        help='Limit number of divisions to scrape')
    parser.add_argument('--max-categories', type=int, default=None,
                        help='Limit number of building product categories to scrape')
    parser.add_argument('--max-companies', type=int, default=None,
                        help='Limit number of companies per division/category')
    parser.add_argument('--output', type=str,
                        default="c:/Users/user1/Desktop/PersonalGit/Scraper/ARCAT_Scraped_Data.xlsx",
                        help='Output Excel file path')
    parser.add_argument('--resume', action='store_true',
                        help='Resume from last checkpoint if available')
    parser.add_argument('--clear-checkpoint', action='store_true',
                        help='Clear existing checkpoint and start fresh')

    args = parser.parse_args()

    scraper = ARCATScraper()

    # Clear checkpoint if requested
    if args.clear_checkpoint:
        scraper.clear_checkpoint()
        print("Checkpoint cleared.")

    print("Starting ARCAT Scraper...")
    print("=" * 50)

    if args.resume:
        print("Resume mode enabled - will continue from last checkpoint if available")

    total_division_companies = 0
    total_category_companies = 0

    try:
        # CSI-Only mode (recommended - clean and fast)
        if args.mode == 'csi-only':
            print("\n--- CSI-ONLY MODE (Recommended) ---")
            print("This scrapes ~28 CSI divisions with ~2,000-5,000 companies total")
            print("Much faster than 'both' mode which has 450K+ companies!\n")
            scrape_csi_only(
                scraper,
                max_divisions=args.max_divisions,
                max_companies=args.max_companies,
                resume=args.resume
            )
            total_division_companies = sum(len(d.companies) for d in scraper.divisions)
            print(f"Divisions scraped: {len(scraper.divisions)}")
            print(f"Companies scraped: {total_division_companies}")

        # Scrape divisions (CSI specs) - legacy mode
        elif args.mode in ['divisions', 'both']:
            print("\n--- Scraping CSI Divisions ---")
            scraper.scrape_all(
                max_divisions=args.max_divisions,
                max_companies_per_division=args.max_companies,
                resume=args.resume
            )
            total_division_companies = sum(len(d.companies) for d in scraper.divisions)
            print(f"Divisions scraped: {len(scraper.divisions)}")
            print(f"Companies from divisions: {total_division_companies}")

        # Scrape building product categories
        if args.mode in ['categories', 'both']:
            print("\n--- Scraping Building Product Categories ---")
            scraper.scrape_building_products_all(
                max_categories=args.max_categories,
                max_subcategories_per_category=None,
                max_companies_per_subcategory=args.max_companies,
                resume=args.resume
            )
            total_category_companies = sum(len(c.companies) for c in scraper.building_product_categories)
            print(f"Categories scraped: {len(scraper.building_product_categories)}")
            print(f"Companies from categories: {total_category_companies}")

        # Export to Excel
        scraper.export_to_excel(args.output)

        # Clear checkpoint on successful completion
        scraper.clear_checkpoint()

        print("\n" + "=" * 50)
        print(f"Scraping complete! Data saved to: {args.output}")
        print(f"Total companies: {total_division_companies + total_category_companies}")

    except KeyboardInterrupt:
        print("\n\nInterrupted by user. Progress has been saved.")
        print(f"Resume with: python arcat_scraper.py --resume --mode {args.mode}")
        # Export whatever we have so far
        scraper.export_to_excel(args.output)
        print(f"Partial data saved to: {args.output}")
    except Exception as e:
        print(f"\n\nError occurred: {e}")
        print("Progress has been saved. Resume with: python arcat_scraper.py --resume")
        # Export whatever we have so far
        try:
            scraper.export_to_excel(args.output)
            print(f"Partial data saved to: {args.output}")
        except:
            pass
        raise


def scrape_csi_only(scraper, max_divisions: int = None, max_companies: int = None, resume: bool = False):
    """
    Scrape CSI divisions from Building Product Categories page.

    Navigation path:
    1. https://www.arcat.com/products/building_products_categories
    2. → Related CSI Divisions (sidebar) e.g., "02 - EXISTING CONDITIONS"
    3. → /content-type/product/existing-conditions-02/existing-conditions-020000
    4. → Company pages e.g., /company/invisible-structures-inc-33364

    Expected output: ~500-2,000 companies total (focused on building product manufacturers)
    """
    logger.info("=" * 60)
    logger.info("STARTING CSI-ONLY SCRAPE (Building Product Categories)")
    logger.info("Source: https://www.arcat.com/products/building_products_categories")
    logger.info("Path: Building Products → Related CSI Divisions → Manufacturers")
    logger.info("=" * 60)

    # Track already scraped companies for resume
    scraped_company_urls = set()

    if resume:
        checkpoint = scraper._load_checkpoint()
        if checkpoint and checkpoint.get('mode') == 'csi-only':
            scraper._restore_from_checkpoint(checkpoint)
            for div in scraper.divisions:
                for company in div.companies:
                    if company.address:
                        scraped_company_urls.add(company.url)
            logger.info(f"Resuming with {len(scraped_company_urls)} already scraped companies")

    try:
        # Step 1: Get Related CSI Divisions from Building Product Categories page
        if not scraper.divisions:
            scraper.scrape_related_csi_divisions()

        logger.info(f"Found {len(scraper.divisions)} Related CSI divisions")

        divisions_to_process = scraper.divisions[:max_divisions] if max_divisions else scraper.divisions

        # First pass: count total companies
        logger.info("Counting companies across all divisions...")
        total_companies = 0
        for division in divisions_to_process:
            if not division.companies:
                scraper.scrape_division_manufacturers(division)
            companies = division.companies[:max_companies] if max_companies else division.companies
            total_companies += len([c for c in companies if c.url not in scraped_company_urls])

        logger.info(f"Total companies to scrape: {total_companies:,}")
        scraper.progress.start(total_companies)

        # Second pass: scrape company details
        for div_idx, division in enumerate(divisions_to_process):
            logger.info(f"\nProcessing division {div_idx + 1}/{len(divisions_to_process)}: {division.code} - {division.name}")

            if scraper.interrupted:
                break

            companies_to_process = division.companies[:max_companies] if max_companies else division.companies

            for company in companies_to_process:
                if company.url in scraped_company_urls:
                    continue

                if scraper.interrupted:
                    break

                scraper.scrape_company_details(company)
                scraper.companies_scraped_count += 1
                scraper.progress.update()

                # Periodic checkpoint
                scraper._maybe_save_checkpoint("csi-only")

                # Progress update every 10 companies
                if scraper.companies_scraped_count % 10 == 0:
                    logger.info(f"  {scraper.progress.get_status_line()}")

            if scraper.interrupted:
                break

        if not scraper.interrupted:
            logger.info(f"\nCSI-Only scraping complete!")
            logger.info(f"Final: {scraper.progress.get_status_line()}")
            scraper._save_checkpoint("csi-only")

        return scraper.divisions

    except Exception as e:
        logger.error(f"Error during scraping: {e}")
        scraper._save_checkpoint("csi-only")
        raise
    finally:
        scraper.close()


def scrape_categories_only():
    """Convenience function to only scrape building product categories"""
    scraper = ARCATScraper()

    print("Starting Building Product Categories Scraper...")
    print("=" * 50)

    scraper.scrape_building_products_all()

    output_file = "c:/Users/user1/Desktop/PersonalGit/Scraper/ARCAT_Scraped_Data.xlsx"
    scraper.export_to_excel(output_file)

    print("=" * 50)
    print(f"Scraping complete! Data saved to: {output_file}")
    total_companies = sum(len(c.companies) for c in scraper.building_product_categories)
    print(f"Total categories scraped: {len(scraper.building_product_categories)}")
    print(f"Total companies found: {total_companies}")


if __name__ == "__main__":
    main()
